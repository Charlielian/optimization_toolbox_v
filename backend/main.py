"""
FastAPI 主入口
- RESTful 接口
- 静态前端挂载
- 全局状态(内存存储)
"""
from __future__ import annotations

import gc
import logging
import os
import re
import sys
import threading
import uuid
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Union

from fastapi import FastAPI, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

# 加入backend到路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app_paths import resolve_resource, runtime_root
from config import config
from log_config import setup_logging
from memory_utils import release_planning_temp
from runtime_utils import (
    plan_lock_busy_detail,
    planning_executor_max_workers,
    release_plan_lock,
    try_acquire_plan_lock,
)
setup_logging()  # 在所有模块 import 之前, 抓取启动期日志
logger = logging.getLogger(__name__)

from license_check import (
    check_license,
    clock_guard_from_config,
    ensure_license_or_exit,
    expired_html,
)

_CLOCK_GUARD = clock_guard_from_config(config.raw)
_LICENSE_STATUS = ensure_license_or_exit(
    config.license_enabled,
    config.license_file,
    clock=_CLOCK_GUARD,
)


from cell_filters import filter_cells_for_map_and_plan
from conflict_check import collect_conflicts, get_directional_skip_count, stats_summary
from cell_sync import sync_cells_from_config
from data_parser import (
    apply_bulk_sector_updates,
    describe_file_rat_profile,
    finalize_manual_cell,
    generate_bulk_sector_update_template,
    parse_bulk_sector_updates,
    parse_work_params,
)
from db import clear_all as db_clear_all
from db import init_db, init_config_db, load_all as db_load_all, save_all as db_save_all
from db import (
    count_cells,
    load_cells_extent,
    load_cells_for_map,
    load_cells_page,
    load_meta_only,
    save_config_sheet,
    get_config_table_data,
    list_config_tables,
    get_import_history,
    delete_config_table,
    get_column_config,
    save_column_config,
    list_sheet_configs,
    delete_sheet_config,
)
from config_parser import parse_config_excel, list_excel_sheets
from config_imports import (
    load_import_config,
    get_enabled_sheets,
    list_all_config_sheets,
    reload_config,
    save_sheet_config_to_yaml,
)
from exporter import (
    export_conflicts,
    export_interference_report,
    export_mml,
    export_neighbors,
    export_plan_split_sheets,
    export_plan_summary,
    export_workparams,
)
from geo_utils import build_sector, set_scene_mode
from interference_analysis import analyze_interference
from nbr_planner import detect_redundancy, plan_neighbors
from pci_planner import plan_all, plan_partial, plan_verify_and_fix
from pci_quality_report import (
    attach_pci_quality_to_cells,
    build_cell_pci_quality,
    build_pci_quality_report,
)
from site_planner import plan_batch_sites, plan_single_site
from app_settings import (
    batch_default_nbr_score_threshold,
    get_plan_defaults,
    neighbor_kwargs_from_defaults,
    pci_defaults_dict,
    reset_plan_defaults,
    save_plan_defaults,
    single_site_default_score_threshold,
)

# 路径（从配置文件读取；与 config.ROOT_DIR 一致）
ROOT_DIR = runtime_root()
config.temp_dir.mkdir(exist_ok=True)
BATCH_RESULT_DIR = config.temp_dir / "batch_results"
BATCH_RESULT_DIR.mkdir(exist_ok=True)


def _cleanup_batch_results(max_age_sec: int = 1800) -> int:
    """清理超过 max_age_sec 秒的批量规划结果文件 (默认30分钟)."""
    import time as _time
    now = _time.time()
    removed = 0
    for p in BATCH_RESULT_DIR.iterdir():
        try:
            if now - p.stat().st_mtime > max_age_sec:
                p.unlink()
                removed += 1
        except Exception:
            continue
    return removed

app = FastAPI(
    title="网优百宝箱",
    description="轻量化4G/5G小区规划工具 - PCI智能规划、加权邻区、冲突检测、MML导出",
    version="1.2.1",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=config.cors_allow_origins,
    allow_methods=config.cors_allow_methods,
    allow_headers=config.cors_allow_headers,
)


@app.middleware("http")
async def _license_gate(request: Request, call_next):
    if not config.license_enabled:
        return await call_next(request)
    status = check_license(
        config.license_enabled,
        config.license_file,
        clock=_CLOCK_GUARD,
    )
    if status.valid:
        return await call_next(request)
    accept = (request.headers.get("accept") or "").lower()
    path = request.url.path
    if path.startswith("/api") or "application/json" in accept:
        return JSONResponse(
            status_code=403,
            content={"detail": status.message},
        )
    return HTMLResponse(expired_html(status.message), status_code=403)


# ──────────────────────────────────────────────
# 日志: 请求 + 异常
# ──────────────────────────────────────────────
@app.middleware("http")
async def _access_log(request: Request, call_next):
    start = datetime.utcnow()
    try:
        response = await call_next(request)
        status = response.status_code
    except Exception:
        logger.exception("未捕获异常 %s %s", request.method, request.url.path)
        raise
    elapsed_ms = (datetime.utcnow() - start).total_seconds() * 1000
    logger.info(
        "%s %s -> %d (%.1fms)",
        request.method, request.url.path, status, elapsed_ms,
    )
    return response


@app.exception_handler(Exception)
async def _global_exception_handler(request: Request, exc: Exception):
    logger.exception("处理请求出错: %s %s", request.method, request.url.path)
    return JSONResponse(
        status_code=500,
        content={"detail": f"{type(exc).__name__}: {exc}"},
    )


@app.exception_handler(RuntimeError)
async def _runtime_error_handler(request: Request, exc: RuntimeError):
    """RuntimeError 视为业务校验失败 (422), 例如 SSS/mod3 约束违反."""
    msg = str(exc)
    # SSS 约束是已知业务校验, 单独标记
    is_sss_violation = "SSS约束" in msg or "mod3约束" in msg
    logger.warning("业务校验失败 [%s %s]: %s", request.method, request.url.path, msg[:300])
    return JSONResponse(
        status_code=422,
        content={
            "detail": msg,
            "error_type": "sss_violation" if is_sss_violation else "runtime_error",
            "hint": (
                "同站小区 N_ID(1) 共享 / mod3 隔离约束违反. "
                "请检查工参 site_name 字段或经纬度聚类是否正确, "
                "以及是否存在异常 PCI 输入."
            ) if is_sss_violation else None,
        },
    )

# 全局状态
class AppState:
    cells: List[Dict[str, Any]] = []
    plan_log: List[str] = []
    plan_stats: Dict[str, Any] = {}

STATE = AppState()

# ── 异步任务队列（用于单站/批量规划超时保护）──
_job_store: Dict[str, Dict[str, Any]] = {}
_executor = ThreadPoolExecutor(max_workers=planning_executor_max_workers())

# 工参：SQLite 为准；规划/写操作前再载入内存，避免启动即占满 RAM
_cells_loaded: bool = False
_temp_cells: List[Dict[str, Any]] = []  # 单站 PLAN-* 等 is_temp，不入库


def _start_job(fn, *args, **kwargs) -> str:
    """提交后台任务，返回 job_id"""
    job_id = uuid.uuid4().hex[:8]
    _job_store[job_id] = {"status": "running", "result": None, "error": None}

    def _run():
        try:
            result = fn(*args, **kwargs)
            _job_store[job_id] = {"status": "done", "result": result, "error": None}
        except Exception as e:
            _job_store[job_id] = {"status": "error", "result": None, "error": str(e)}

    threading.Thread(target=_run, daemon=True).start()
    return job_id


@app.get("/api/job/{job_id}")
async def api_job_status(job_id: str):
    """查询异步任务状态（前端轮询用）"""
    job = _job_store.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="任务不存在或已过期")
    return {"job_id": job_id, **job}


def _persist_current() -> None:
    """把当前 STATE.cells + 规划统计 整体入库"""
    try:
        db_save_all(STATE.cells, {
            "stats": STATE.plan_stats,
            "cells_count": len(STATE.cells),
            "last_saved_at": datetime.utcnow().isoformat(timespec="seconds"),
        })
    except Exception as e:
        print(f"[warn] 持久化失败: {e}", flush=True)


def _refresh_workparam_stats() -> None:
    """工参增删改后刷新内存统计（制式分布等）。"""
    from collections import Counter

    cells = STATE.cells
    rc = Counter(c.get("rat", "LTE") for c in cells)
    STATE.plan_stats = {
        **(STATE.plan_stats or {}),
        "total": len(cells),
        "valid": len(cells),
        "rat_counts": dict(rc),
        "cells_count": len(cells),
    }


def _find_cell_index(ecgi: str) -> int:
    for i, c in enumerate(STATE.cells):
        if c.get("ecgi") == ecgi:
            return i
    return -1


def _cell_to_light(c: Dict[str, Any], *, include_neighbors: bool = True) -> Dict[str, Any]:
    nbrs = c.get("neighbors", []) if include_neighbors else []
    out = {
        "ecgi": c.get("ecgi"),
        "name": c.get("name"),
        "rat": c.get("rat"),
        "earfcn": c.get("earfcn"),
        "lon": c.get("lon"),
        "lat": c.get("lat"),
        "azimuth": c.get("azimuth"),
        "radius": c.get("radius"),
        "beamwidth": c.get("beamwidth", c.get("beam")),
        "beam": c.get("beam"),
        "freq_band": c.get("freq_band"),
        "freq_band_label": c.get("freq_band_label"),
        "site_type": c.get("site_type", "陆地"),
        "site_type_label": c.get("site_type_label"),
        "tac": c.get("tac"),
        "pci": c.get("pci"),
        "new_pci": c.get("new_pci", c.get("pci")),
        "site_name": c.get("site_name"),
        "phy_name": c.get("phy_name"),
        "ant_name": c.get("ant_name"),
        "manufacturer": c.get("manufacturer"),
        "oms_name": c.get("oms_name"),
        "cell_id": c.get("cell_id"),
        "bandwidth": c.get("bandwidth"),
        "freq_band_ind": c.get("freq_band_ind"),
        "pci_synced_at": c.get("pci_synced_at"),
        "neighbor_count": len(nbrs),
        "pci_quality": c.get("pci_quality"),
    }
    if include_neighbors:
        out["neighbors"] = nbrs
    return out


def _merge_runtime_cells() -> List[Dict[str, Any]]:
    """持久化工参 + 内存临时小区（单站 PLAN-*）。"""
    if not _temp_cells:
        return STATE.cells
    return STATE.cells + _temp_cells


def _ensure_cells_loaded() -> None:
    """按需从 SQLite 载入工参到 STATE（启动不预加载）。"""
    global _cells_loaded
    if _cells_loaded:
        return
    cells, meta = db_load_all()
    STATE.cells = cells or []
    if meta.get("stats"):
        STATE.plan_stats = {**(STATE.plan_stats or {}), **meta["stats"]}
    _cells_loaded = True
    logger.info("从数据库载入 %d 个小区到内存", len(STATE.cells))


def _require_cells_loaded() -> None:
    _ensure_cells_loaded()
    if not STATE.cells and not _temp_cells:
        raise HTTPException(400, "请先上传工参文件")


def _plan_busy_response() -> JSONResponse:
    return JSONResponse(
        status_code=409,
        content={
            "detail": f"已有规划任务进行中（{plan_lock_busy_detail()}），请稍后再试",
        },
    )


@app.on_event("startup")
def _on_startup():
    """启动时: 建表；工参留 SQLite，按需再载入内存"""
    init_db()
    init_config_db()  # 初始化配置相关表
    meta = load_meta_only()
    STATE.plan_stats = meta.get("stats") or {}
    n = count_cells()
    if n:
        logger.info("数据库已有 %d 个小区，将在首次规划/写操作前载入内存", n)
    else:
        logger.info("数据库为空, 等待上传工参")


# =========================
# Pydantic 模型
# =========================
class PlanAllRequest(BaseModel):
    """全网规划：默认仅 PCI；plan_neighbors=True 时追加邻区规划。"""
    plan_neighbors: bool = False
    max_neighbors: int = Field(default=16, ge=1, le=800)
    max_distance_km: float = Field(default=5.0, gt=0)
    min_overlap_ratio: float = Field(default=0.0, ge=0)
    score_threshold: float = Field(default=0.10, ge=0, le=1)
    enable_cross_system: bool = True
    enable_bidirectional: bool = True
    weight_distance: float = Field(default=0.7, ge=0, le=1)
    weight_overlap: float = Field(default=0.3, ge=0, le=1)
    scene_mode: str = "land"
    directional_filter: bool = True
    # PCI 引擎参数（与 PCI 大屏 / 单站规划一致）
    engine: str = Field(default="legacy", pattern="^(legacy|rftools)$")
    reuse_distance_km: float = Field(default=5.0, ge=0.1, le=100.0)
    check_mod6: bool = False
    check_mod30: bool = True
    pci_whitelist: Optional[List[int]] = None
    pci_blacklist: Optional[List[int]] = None
    rat: Optional[str] = Field(default=None, description="LTE=4G, NR=5G, 空=全部")
    freq_band: Optional[str] = Field(default=None, description="频段标签, 空=全部")


class PlanPartialRequest(BaseModel):
    selected_ecgis: List[str]
    radius_km: float = Field(default=5.0, gt=0)
    plan_neighbors: bool = False
    max_neighbors: int = 16
    max_distance_km: float = 5.0
    enable_cross_system: bool = True
    directional_filter: bool = True
    engine: str = Field(default="legacy", pattern="^(legacy|rftools)$")
    reuse_distance_km: float = Field(default=5.0, ge=0.1, le=100.0)
    check_mod6: bool = False
    check_mod30: bool = True
    pci_whitelist: Optional[List[int]] = None
    pci_blacklist: Optional[List[int]] = None
    rat: Optional[str] = None
    freq_band: Optional[str] = None


class ExportRequest(BaseModel):
    export_type: str = Field(default="workparams", pattern="^(workparams|conflicts|neighbors|mml|summary)$")
    vendor: str = "huawei"


class CellCreateRequest(BaseModel):
    """界面手工新增工参小区"""
    ecgi: str
    name: str
    rat: str
    lon: float
    lat: float
    azimuth: float = 0.0
    pci: Optional[int] = None
    earfcn: Optional[int] = None
    tac: Optional[int] = None
    site_type: str = "陆地"
    phy_name: Optional[str] = None
    ant_name: Optional[str] = None
    manufacturer: Optional[str] = None
    oms_name: Optional[str] = None
    freq_band_raw: Optional[str] = None
    bandwidth: Optional[float] = None


class CellUpdateRequest(BaseModel):
    """界面编辑工参小区（不可改 ECGI）"""
    name: Optional[str] = None
    rat: Optional[str] = None
    lon: Optional[float] = None
    lat: Optional[float] = None
    azimuth: Optional[float] = None
    pci: Optional[int] = None
    earfcn: Optional[int] = None
    tac: Optional[int] = None
    site_type: Optional[str] = None
    phy_name: Optional[str] = None
    ant_name: Optional[str] = None
    manufacturer: Optional[str] = None
    oms_name: Optional[str] = None
    freq_band_raw: Optional[str] = None
    bandwidth: Optional[float] = None


# =========================
# 接口
# =========================
@app.post("/api/clear")
async def api_clear():
    """清空数据库 + 内存中的所有小区和规划结果"""
    global _cells_loaded
    db_clear_all()
    STATE.cells = []
    STATE.plan_log = []
    STATE.plan_stats = {}
    _temp_cells.clear()
    _cells_loaded = True
    gc.collect()
    return {
        "success": True,
        "cells_count": 0,
        "message": "数据库已清空",
    }


def _apply_workparam_import(
    new_cells: List[Dict[str, Any]],
    stats: Dict[str, Any],
    invalid_rows: List[Dict[str, Any]],
    append: bool,
) -> Dict[str, Any]:
    """单批工参写入 STATE，返回与 /api/upload 一致的响应字段。"""
    from collections import Counter

    global _cells_loaded
    _ensure_cells_loaded()

    if append and STATE.cells:
        existing = {c["ecgi"]: c for c in STATE.cells if c.get("ecgi")}
        added, updated = 0, 0
        for c in new_cells:
            ecgi = c.get("ecgi")
            if not ecgi:
                added += 1
                continue
            if ecgi in existing:
                updated += 1
            else:
                added += 1
            existing[ecgi] = c
        merged = list(existing.values())
        STATE.cells = merged
        STATE.plan_log = []
        rc = Counter(c.get("rat", "LTE") for c in merged)
        STATE.plan_stats = {
            "total": len(merged),
            "valid": len(merged),
            "invalid": stats.get("invalid", 0),
            "rat_counts": dict(rc),
        }
        _persist_current()
        return {
            "success": True,
            "mode": "append",
            "added": added,
            "updated": updated,
            "kept": 0,
            "total_after": len(merged),
            "stats": stats,
            "invalid_rows": invalid_rows,
            "cells_count": len(merged),
        }

    STATE.cells = list(new_cells)
    STATE.plan_log = []
    _cells_loaded = True
    STATE.plan_stats = {
        "total": stats.get("total", len(new_cells)),
        "valid": stats.get("valid", len(new_cells)),
        "invalid": stats.get("invalid", 0),
        "rat_counts": stats.get("rat_counts") or {},
    }
    _persist_current()
    band_dist = Counter(c.get("freq_band_label") for c in STATE.cells)
    sector_dist = Counter((c.get("beam"), c.get("radius")) for c in STATE.cells)
    out = {
        "success": True,
        "mode": "replace",
        "stats": stats,
        "invalid_rows": invalid_rows,
        "cells_count": len(STATE.cells),
        "band_distribution": dict(band_dist.most_common()),
        "sector_distribution": {f"{b},{r}": n for (b, r), n in sector_dist.most_common()},
    }
    del new_cells
    gc.collect()
    return out


@app.post("/api/upload")
async def api_upload(
    file: UploadFile = File(...),
    append: bool = Form(False),
):
    """上传工参文件(Excel/CSV)

    append=False (默认): 替换现有工参 (适合同制式单文件导入)
    append=True:         追加到现有工参 (适合 4G/5G 分别上传后再合并)
    """
    content = await file.read()
    if not content:
        raise HTTPException(400, "文件为空")
    result = parse_work_params(content, file.filename or "upload.xlsx")
    if "error" in result and not result.get("valid_cells"):
        raise HTTPException(400, result["error"])

    return _apply_workparam_import(
        result["valid_cells"],
        result["stats"],
        result.get("invalid_rows") or [],
        append,
    )


@app.post("/api/upload/batch")
async def api_upload_batch(
    files: List[UploadFile] = File(...),
    append: bool = Form(False),
):
    """多文件工参导入：依次解析并合并，按 ECGI 去重；返回每个文件的 4G/5G 识别结果。"""
    if not files:
        raise HTTPException(400, "请选择至少一个文件")

    file_reports: List[Dict[str, Any]] = []
    merged_cells: Dict[str, Dict[str, Any]] = {}
    total_stats = {"total": 0, "valid": 0, "invalid": 0, "rat_counts": {"LTE": 0, "NR": 0}}
    all_invalid: List[Dict[str, Any]] = []

    for uf in files:
        fname = uf.filename or "upload.xlsx"
        content = await uf.read()
        if not content:
            file_reports.append({
                "filename": fname,
                "success": False,
                "error": "文件为空",
                "rat_profile": describe_file_rat_profile(fname, {}, 0),
            })
            continue

        result = parse_work_params(content, fname)
        st = result.get("stats") or {}
        profile = describe_file_rat_profile(
            fname, st.get("rat_counts") or {}, st.get("valid", 0)
        )

        if "error" in result and not result.get("valid_cells"):
            file_reports.append({
                "filename": fname,
                "success": False,
                "error": result.get("error", "解析失败"),
                "stats": st,
                "rat_profile": profile,
                "invalid_rows": result.get("invalid_rows") or [],
            })
            continue

        for c in result.get("valid_cells") or []:
            ecgi = c.get("ecgi")
            if ecgi:
                merged_cells[ecgi] = c

        total_stats["total"] += st.get("total", 0)
        total_stats["valid"] += st.get("valid", 0)
        total_stats["invalid"] += st.get("invalid", 0)
        for rat, n in (st.get("rat_counts") or {}).items():
            total_stats["rat_counts"][rat] = total_stats["rat_counts"].get(rat, 0) + n

        inv = result.get("invalid_rows") or []
        for row in inv:
            row = dict(row)
            row["source_file"] = fname
            all_invalid.append(row)

        file_reports.append({
            "filename": fname,
            "success": True,
            "stats": st,
            "rat_profile": profile,
            "invalid_rows": inv,
        })

    if not merged_cells and not any(r.get("success") for r in file_reports):
        raise HTTPException(400, "所有文件均无法导入")

    new_list = list(merged_cells.values())
    total_stats["rat_counts"] = {
        "LTE": sum(1 for c in new_list if c.get("rat") == "LTE"),
        "NR": sum(1 for c in new_list if c.get("rat") == "NR"),
    }
    total_stats["valid"] = len(new_list)

    resp = _apply_workparam_import(new_list, total_stats, all_invalid, append)
    resp["file_count"] = len(files)
    resp["files"] = file_reports
    return resp


@app.post("/api/workparams/bulk-update")
async def api_workparams_bulk_update(file: UploadFile = File(...)):
    """
    按 CGI/ECGI 关联批量更新工参中的 pci、tac、earfcn（csv/xlsx/xls）。
    文件需含 cgi 列及 pci/tac/earfcn 中至少一列有值；空单元格表示不修改该字段。
    """
    _require_cells_loaded()
    content = await file.read()
    if not content:
        raise HTTPException(400, "文件为空")

    parsed = parse_bulk_sector_updates(content, file.filename or "update.xlsx")
    if not parsed.get("success"):
        raise HTTPException(400, parsed.get("error", "解析失败"))

    rows = parsed.get("rows") or []
    if not rows and not parsed.get("invalid_rows"):
        raise HTTPException(400, "文件无有效更新行")

    apply_result = apply_bulk_sector_updates(STATE.cells, rows)
    if apply_result.get("updated", 0) > 0:
        _refresh_workparam_stats()
        _persist_current()

    return {
        "success": True,
        "filename": file.filename,
        "parse_stats": parsed.get("stats"),
        "invalid_rows": (parsed.get("invalid_rows") or [])[:100],
        **apply_result,
        "cells_count": len(STATE.cells),
    }


@app.get("/api/workparams/bulk-update/template")
async def api_workparams_bulk_update_template():
    """下载 CGI + pci/tac/earfcn 四列更新模板。"""
    from urllib.parse import quote

    blob = generate_bulk_sector_update_template()
    ascii_name = "sector_update_template.xlsx"
    cn_name = "工参更新模板_cgi_pci_tac_earfcn.xlsx"
    headers = {
        "Content-Disposition": (
            f'attachment; filename="{ascii_name}"; '
            f"filename*=UTF-8''{quote(cn_name)}"
        ),
    }
    return Response(
        blob,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.post("/api/plan/all")
async def api_plan_all(req: PlanAllRequest):
    """全网 PCI 规划；可选邻区规划 (plan_neighbors=True)"""
    _require_cells_loaded()
    if not try_acquire_plan_lock("plan_all"):
        return _plan_busy_response()

    try:
        set_scene_mode(req.scene_mode)

        from site_type_ext import build_per_site_thresholds
        per_site = build_per_site_thresholds(
            STATE.cells, default_safe_m=1500.0, default_same_pci_min_m=30000.0,
        )

        pci_result = plan_all(
            STATE.cells,
            whitelist=req.pci_whitelist,
            blacklist=req.pci_blacklist,
            engine=req.engine,
            reuse_distance_km=req.reuse_distance_km,
            check_mod6=req.check_mod6,
            check_mod30=req.check_mod30,
            per_site_thresholds=per_site,
            directional_filter=req.directional_filter,
            rat_filter=req.rat,
            freq_band_filter=req.freq_band,
        )
        STATE.plan_log.extend(pci_result["log"])
        per_site = None

        nbr_stats: Dict[str, Any] = {}
        nbr_result = None
        if req.plan_neighbors:
            nbr_result = plan_neighbors(
                STATE.cells,
                max_neighbors=req.max_neighbors,
                max_distance_km=req.max_distance_km,
                min_overlap_ratio=req.min_overlap_ratio,
                enable_cross_system=req.enable_cross_system,
                enable_bidirectional=req.enable_bidirectional,
                score_threshold=req.score_threshold,
                weight_distance=req.weight_distance,
                weight_overlap=req.weight_overlap,
            )
            STATE.plan_log.extend(nbr_result["log"])
            nbr_stats = nbr_result["stats"]

        conflicts = collect_conflicts(
            STATE.cells, use_original_pci=False, directional_filter=req.directional_filter,
        )
        STATE.plan_stats.update({
            "conflict_count": len(conflicts),
            "conflict_stats": stats_summary(conflicts),
            "pci_stats": pci_result["stats"],
        })
        if nbr_stats:
            STATE.plan_stats["neighbor_stats"] = nbr_stats
        pci_quality = _pci_quality_after_plan(
            conflicts,
            check_mod30=req.check_mod30,
            directional_filter=req.directional_filter,
        )
        attach_pci_quality_to_cells(STATE.cells, pci_quality)
        STATE.plan_stats["pci_quality_summary"] = pci_quality.get("summary", {})
        _persist_current()

        resp = {
            "success": True,
            "stats": STATE.plan_stats,
            "log": pci_result.get("log", STATE.plan_log[-50:]),
            "conflicts_count": len(conflicts),
            "conflicts": conflicts[:200],
            "pci_quality": pci_quality,
        }
        release_planning_temp(trim_log=STATE.plan_log)
        return resp
    finally:
        release_plan_lock()


@app.post("/api/plan/partial")
async def api_plan_partial(req: PlanPartialRequest):
    """局部 PCI 微调；可选邻区规划"""
    _require_cells_loaded()
    if not try_acquire_plan_lock("plan_partial"):
        return _plan_busy_response()

    try:
        return _api_plan_partial_body(req)
    finally:
        release_plan_lock()


def _api_plan_partial_body(req: PlanPartialRequest) -> Dict[str, Any]:
    per_site = None
    nbr_result = None
    if req.engine == "rftools":
        # 局部模式：仅对选中小区分配 PCI
        idx_map = {c["ecgi"]: i for i, c in enumerate(STATE.cells)}
        target_indices = [idx_map[e] for e in req.selected_ecgis if e in idx_map]
        if not target_indices:
            raise HTTPException(400, "selected_ecgis 在工参中不存在")
        from site_type_ext import build_per_site_thresholds
        per_site = build_per_site_thresholds(STATE.cells, default_safe_m=1500.0, default_same_pci_min_m=30000.0)
        pci_result = plan_all(
            STATE.cells,
            whitelist=req.pci_whitelist,
            blacklist=req.pci_blacklist,
            engine="rftools",
            reuse_distance_km=req.reuse_distance_km,
            check_mod6=req.check_mod6,
            check_mod30=req.check_mod30,
            per_site_thresholds=per_site,
            target_indices=target_indices,
            directional_filter=req.directional_filter,
            rat_filter=req.rat,
            freq_band_filter=req.freq_band,
        )
    else:
        from site_type_ext import build_per_site_thresholds
        per_site = build_per_site_thresholds(STATE.cells, default_safe_m=1500.0, default_same_pci_min_m=30000.0)
        pci_result = plan_partial(
            STATE.cells,
            req.selected_ecgis,
            req.radius_km,
            whitelist=req.pci_whitelist,
            blacklist=req.pci_blacklist,
            directional_filter=req.directional_filter,
            per_site_thresholds=per_site,
            check_mod30=req.check_mod30,
            rat_filter=req.rat,
            freq_band_filter=req.freq_band,
        )
    STATE.plan_log.extend(pci_result["log"])

    nbr_stats: Dict[str, Any] = {}
    if req.plan_neighbors:
        nbr_kw = neighbor_kwargs_from_defaults(
            {
                "max_neighbors": req.max_neighbors,
                "max_distance_km": req.max_distance_km,
                "enable_cross_system": req.enable_cross_system,
            }
        )
        nbr_result = plan_neighbors(STATE.cells, **nbr_kw)
        STATE.plan_log.extend(nbr_result.get("log", []))
        nbr_stats = nbr_result["stats"]

    conflicts = collect_conflicts(STATE.cells, use_original_pci=False,
                                   directional_filter=req.directional_filter)
    STATE.plan_stats.update({
        "conflict_count": len(conflicts),
        "conflict_stats": stats_summary(conflicts),
        "pci_stats": pci_result.get("stats", {}),
    })
    if nbr_stats:
        STATE.plan_stats["neighbor_stats"] = nbr_stats
    affected_set = set(pci_result.get("affected") or req.selected_ecgis)
    pci_quality = _pci_quality_after_plan(
        conflicts,
        check_mod30=req.check_mod30,
        directional_filter=req.directional_filter,
        ecgi_filter=affected_set if affected_set else None,
    )
    attach_pci_quality_to_cells(STATE.cells, pci_quality)
    STATE.plan_stats["pci_quality_summary"] = pci_quality.get("summary", {})
    _persist_current()

    resp = {
        "success": True,
        "affected": pci_result.get("affected", []),
        "conflicts_count": len(conflicts),
        "conflicts": conflicts[:200],
        "log": pci_result["log"],
        "stats": STATE.plan_stats,
        "pci_quality": pci_quality,
    }
    release_planning_temp(trim_log=STATE.plan_log)
    return resp


class PciQualityRequest(BaseModel):
    ecgis: Optional[List[str]] = None
    check_mod30: bool = True
    directional_filter: bool = True


@app.post("/api/pci/quality")
async def api_pci_quality(req: PciQualityRequest):
    """基于当前 new_pci 生成 PCI 得分、最近邻与干扰说明（规划后或手动改 PCI 后均可调用）。"""
    _require_cells_loaded()
    ecgi_set = set(req.ecgis) if req.ecgis else None
    conflicts = collect_conflicts(
        STATE.cells, use_original_pci=False, directional_filter=req.directional_filter,
    )
    report = build_pci_quality_report(
        STATE.cells,
        conflicts,
        ecgi_filter=ecgi_set,
        check_mod30=req.check_mod30,
        directional_filter=req.directional_filter,
    )
    return {"success": True, "pci_quality": report, "conflicts_count": len(conflicts)}


@app.post("/api/check/conflict")
async def api_check_conflict(directional_filter: bool = True):
    """冲突校验"""
    _require_cells_loaded()
    conflicts = collect_conflicts(STATE.cells, use_original_pci=False,
                                   directional_filter=directional_filter)
    return {
        "success": True,
        "conflicts": conflicts,
        "stats": stats_summary(conflicts),
        "directional_filter": directional_filter,
        "directional_skip_count": get_directional_skip_count(),
    }


@app.post("/api/check/redundancy")
async def api_check_redundancy():
    """邻区冗余/漏配检测"""
    _require_cells_loaded()
    result = detect_redundancy(STATE.cells)
    return {
        "success": True,
        **result,
    }


@app.post("/api/export/file")
async def api_export_file(req: ExportRequest):
    """导出文件(工参/报表/MML)"""
    _require_cells_loaded()
    cells_export = _merge_runtime_cells()

    if req.export_type == "workparams":
        content = export_workparams(cells_export)
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=workparams_planned.xlsx"},
        )
    elif req.export_type == "neighbors":
        content = export_neighbors(cells_export)
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=neighbors_list.xlsx"},
        )
    elif req.export_type == "conflicts":
        conflicts = collect_conflicts(cells_export, use_original_pci=False)
        content = export_conflicts(conflicts)
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=conflicts_report.xlsx"},
        )
    elif req.export_type == "mml":
        mml_text = export_mml(cells_export, vendor=req.vendor)
        ext = "txt" if req.vendor == "zte" else "txt"
        return Response(
            content=mml_text,
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=neighbors_{req.vendor}.{ext}"},
        )
    elif req.export_type == "summary":
        conflicts = collect_conflicts(STATE.cells, use_original_pci=False)
        content = export_plan_summary(cells_export, conflicts, STATE.plan_log)
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=plan_summary.xlsx"},
        )
    else:
        raise HTTPException(400, f"不支持的导出类型: {req.export_type}")


@app.get("/api/cells/extent")
async def api_cells_extent(rat: Optional[str] = Query(default=None)):
    """工参经纬度包围盒（规划页首次定位，不拉全量）。"""
    ext = load_cells_extent(rat=rat)
    meta = load_meta_only()
    base_stats = {**(STATE.plan_stats or {}), **(meta.get("stats") or {})}
    n = count_cells()
    return {
        "success": True,
        "extent": ext,
        "stats": {**base_stats, "cells_count": n, "total": base_stats.get("total") or n},
    }


@app.get("/api/cells")
async def api_cells(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    keyword: Optional[str] = Query(default=None),
    rat: Optional[str] = Query(default=None),
    sync_filter: Optional[str] = Query(default=None, pattern="^(synced|unsynced)$"),
    min_lat: Optional[float] = Query(default=None),
    max_lat: Optional[float] = Query(default=None),
    min_lon: Optional[float] = Query(default=None),
    max_lon: Optional[float] = Query(default=None),
    map_limit: int = Query(default=8000, ge=100, le=20000),
    mode: str = Query(default="auto", pattern="^(auto|paged|map|full)$"),
):
    """
    工参查询：默认优先从 SQLite 分页/视口加载，减轻内存与响应体积。
    - mode=paged：工参表分页（不含 neighbors）
    - mode=map 或带 min_lat/max_lat/min_lon/max_lon：地图视口
    - mode=full：载入内存后返回全量（兼容旧前端，慎用大数据集）
    """
    meta = load_meta_only()
    base_stats = {**(STATE.plan_stats or {}), **(meta.get("stats") or {})}
    base_stats["cells_count"] = count_cells()
    base_stats["total"] = base_stats.get("total") or base_stats["cells_count"]

    if mode == "full":
        _ensure_cells_loaded()
        merged = _merge_runtime_cells()
        cells_light = [
            _cell_to_light(c, include_neighbors=False)
            for c in filter_cells_for_map_and_plan(merged)
        ]
        conflicts = collect_conflicts(merged, use_original_pci=False)
        stats = {**base_stats, "conflict_count": len(conflicts)}
        return {
            "success": True,
            "cells": cells_light,
            "stats": stats,
            "truncated": False,
        }

    use_map = mode == "map" or (
        mode == "auto"
        and min_lat is not None
        and max_lat is not None
        and min_lon is not None
        and max_lon is not None
    )

    if not use_map:
        offset = (page - 1) * page_size
        cells, total = load_cells_page(
            offset,
            page_size,
            rat=rat,
            keyword=keyword,
            sync_filter=sync_filter,
            light=True,
        )
        cells_light = [_cell_to_light(c, include_neighbors=False) for c in cells]
        stats = {**base_stats, "total": total, "conflict_count": base_stats.get("conflict_count")}
        return {
            "success": True,
            "cells": cells_light,
            "stats": stats,
            "page": page,
            "page_size": page_size,
            "total": total,
            "truncated": False,
        }

    cells = load_cells_for_map(
        min_lat=min_lat,
        max_lat=max_lat,
        min_lon=min_lon,
        max_lon=max_lon,
        rat=rat,
        limit=map_limit,
    )
    cells = filter_cells_for_map_and_plan(cells)
    cells_light = [_cell_to_light(c, include_neighbors=False) for c in cells]
    truncated = len(cells) >= map_limit
    stats = {**base_stats, "conflict_count": base_stats.get("conflict_count")}
    return {
        "success": True,
        "cells": cells_light,
        "stats": stats,
        "truncated": truncated,
        "map_limit": map_limit,
    }


@app.post("/api/cells")
async def api_cell_create(req: CellCreateRequest):
    """手工新增一条工参小区"""
    from data_parser import RAT_ALIASES

    _ensure_cells_loaded()
    ecgi = (req.ecgi or "").strip()
    if not ecgi:
        raise HTTPException(400, "ECGI 不能为空")
    if _find_cell_index(ecgi) >= 0:
        raise HTTPException(409, f"ECGI 已存在: {ecgi}")

    rat_raw = (req.rat or "").strip()
    rat = RAT_ALIASES.get(rat_raw.upper(), RAT_ALIASES.get(rat_raw))
    if rat not in ("LTE", "NR"):
        raise HTTPException(400, f"未知制式: {req.rat}")

    payload = req.model_dump()
    payload["ecgi"] = ecgi
    payload["rat"] = rat
    cell, errs = finalize_manual_cell(payload)
    if errs:
        raise HTTPException(400, {"detail": "校验失败", "errors": errs})

    cell["updated_at"] = datetime.utcnow().isoformat(timespec="seconds")
    STATE.cells.append(cell)
    _refresh_workparam_stats()
    _persist_current()
    return {"success": True, "cell": _cell_to_light(cell)}


@app.put("/api/cells/{ecgi:path}")
async def api_cell_update(ecgi: str, req: CellUpdateRequest):
    """编辑工参小区（ECGI 不可修改）"""
    from data_parser import RAT_ALIASES

    _ensure_cells_loaded()
    idx = _find_cell_index(ecgi)
    if idx < 0:
        raise HTTPException(404, f"未找到小区: {ecgi}")

    cell = dict(STATE.cells[idx])
    updates = req.model_dump(exclude_unset=True)
    if "rat" in updates and updates["rat"] is not None:
        rat_raw = str(updates["rat"]).strip()
        rat = RAT_ALIASES.get(rat_raw.upper(), RAT_ALIASES.get(rat_raw))
        if rat not in ("LTE", "NR"):
            raise HTTPException(400, f"未知制式: {updates['rat']}")
        updates["rat"] = rat
    for k, v in updates.items():
        cell[k] = v

    cell, errs = finalize_manual_cell(cell)
    if errs:
        raise HTTPException(400, {"detail": "校验失败", "errors": errs})

    cell["updated_at"] = datetime.utcnow().isoformat(timespec="seconds")
    STATE.cells[idx] = cell
    _refresh_workparam_stats()
    _persist_current()
    return {"success": True, "cell": _cell_to_light(cell)}


@app.delete("/api/cells/{ecgi:path}")
async def api_cell_delete(ecgi: str):
    """删除一条工参小区"""
    _ensure_cells_loaded()
    idx = _find_cell_index(ecgi)
    if idx < 0:
        raise HTTPException(404, f"未找到小区: {ecgi}")

    removed = STATE.cells.pop(idx)
    for c in STATE.cells:
        nbrs = c.get("neighbors") or []
        if not nbrs:
            continue
        c["neighbors"] = [n for n in nbrs if n.get("dst_ecgi") != ecgi]

    _refresh_workparam_stats()
    _persist_current()
    return {"success": True, "ecgi": ecgi, "name": removed.get("name")}


@app.get("/api/sample-data")
async def api_sample_data():
    """下载示例工参文件"""
    sample_path = ROOT_DIR / "static" / "sample_cells.xlsx"
    if not sample_path.exists():
        raise HTTPException(404, "示例文件不存在")
    return FileResponse(
        sample_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="sample_cells.xlsx",
    )


@app.get("/api/template")
async def api_template(rat: str = Query("both", pattern="^(4G|5G|both)$")):
    """下载工参模板 (4G / 5G / 双制式)"""
    from urllib.parse import quote
    from template_generator import generate_template

    blob = generate_template(rat)
    ascii_name = {
        "4G": "cell_template_4G.xlsx",
        "5G": "cell_template_5G.xlsx",
        "both": "cell_template_4G_5G.xlsx",
    }[rat]
    # 用 RFC 5987 编码中文 filename* (兼容主流浏览器)
    cn_name = {
        "4G": "工参模板_4G.xlsx",
        "5G": "工参模板_5G.xlsx",
        "both": "工参模板_4G_5G.xlsx",
    }[rat]
    headers = {
        "Content-Disposition": (
            f'attachment; filename="{ascii_name}"; '
            f"filename*=UTF-8''{quote(cn_name)}"
        ),
    }
    return Response(
        blob,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/api/health")
async def api_health():
    return {
        "status": "ok",
        "cells_in_db": count_cells(),
        "cells_in_memory": len(STATE.cells) + len(_temp_cells),
        "memory_loaded": _cells_loaded,
        "version": "1.2.1",
    }


class PlanDefaultsPatch(BaseModel):
    """部分更新规划默认参数（写入 SQLite meta，重启后仍有效）"""
    pci: Optional[Dict[str, Any]] = None
    neighbor: Optional[Dict[str, Any]] = None
    batch: Optional[Dict[str, Any]] = None


@app.get("/api/settings/plan-defaults")
async def api_get_plan_defaults():
    return {"success": True, "defaults": get_plan_defaults()}


@app.put("/api/settings/plan-defaults")
async def api_put_plan_defaults(body: PlanDefaultsPatch):
    try:
        merged = save_plan_defaults(body.model_dump(exclude_none=True))
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"success": True, "defaults": merged}


@app.post("/api/settings/plan-defaults/reset")
async def api_reset_plan_defaults():
    return {"success": True, "defaults": reset_plan_defaults()}


# =========================
# 单站 / 批量规划 (新功能)
# =========================
class SingleSitePlanRequest(BaseModel):
    lat: float
    lon: float
    rat: str = Field(pattern="^(4G|5G|LTE|NR)$", default="5G")
    freq_band: str = "默认"
    plan_site_type: str = Field(pattern="^(macro|micro|indoor)$", default="macro")
    n_sectors: int = Field(ge=1, le=6, default=3)
    base_azimuth: Union[float, List[float]] = Field(default=0)
    nbr_plan_types: List[str] = Field(default_factory=lambda: ["4G_4G", "4G_5G", "5G_4G", "5G_5G"])
    engine: str = Field(pattern="^(legacy|rftools)$", default="legacy")
    reuse_distance_km: float = Field(ge=0.1, le=100.0, default=5.0)
    check_mod6: bool = False
    check_mod30: bool = True
    use_beam_overlap_score: bool = False
    directional_filter: bool = True
    name_hint: Optional[str] = None
    site_name: Optional[str] = None
    earfcn: Optional[int] = None
    tac: Optional[int] = None
    persist: bool = True  # 是否持久化到 DB (默认 True)
    # ── 邻区得分阈值: 低于此得分的候选邻区直接丢弃, 避免冗余邻区 ──
    # 工程经验: 0.1 会包含几乎所有近距离候选 (造成冗余),
    #          0.5 过滤掉意义不大的邻区关系 (推荐默认),
    #          0.7+ 仅保留强相关邻区
    score_threshold: Optional[float] = Field(default=None, ge=0.0, le=1.0)
    # 规划模式: pci - 仅PCI规划, nbr - 仅邻区规划(需已有PCI), pci+nbr - PCI+邻区规划(默认)
    planning_mode: Optional[str] = Field(default=None, pattern="^(pci|nbr|pci\\+nbr)$")


class ExportSplitRequest(BaseModel):
    planned_ecgis: List[str]
    nbr_plan_types: Optional[List[str]] = None


class InterferenceRequest(BaseModel):
    interference_distance_km: float = Field(ge=0.1, le=100.0, default=5.0)
    overlap_threshold: float = Field(ge=0.0, le=100.0, default=30.0)
    detect_co_channel: bool = True
    detect_adjacent_channel: bool = True
    detect_pci_collision: bool = True
    detect_mod3: bool = True
    detect_mod6: bool = False
    center_ecgi: Optional[str] = None
    radius_km: Optional[float] = None  # 只分析此半径内
    # ── 4G/5G + 频段 + 区域圈选 ──
    rat: Optional[str] = None             # "LTE" | "NR" | None
    freq_band: Optional[str] = None       # 频段(空=全频段)
    area: Optional[Dict[str, Any]] = Field(default=None, description="""
        圈选区域 (前端画完图形后传入):
          {type:'rect',  lat1, lon1, lat2, lon2}
          {type:'circle',lat,  lon,  radius_km}
          {type:'polygon', points: [[lat, lon], ...]}
    """)


# ──────────────────────────────────────────────
# SSE 流式响应辅助
# ──────────────────────────────────────────────
def _sse_format(event: str, data: Any) -> bytes:
    """格式化为 SSE 单个事件 (event: ...\\ndata: ...\\n\\n)"""
    import json as _json
    if not isinstance(data, str):
        data = _json.dumps(data, ensure_ascii=False, default=str)
    return f"event: {event}\ndata: {data}\n\n".encode("utf-8")


def _sse_stream_progress(gen_fn, *, loop=None, executor=None, timeout: float = 180.0):
    """
    把"同步生成器(进度回调)"包成 StreamingResponse。
    gen_fn(progress_cb, set_done) -> 最终结果dict
    progress_cb(pct, stage) 应立即把 (pct, stage) 推入队列.
    """
    import asyncio
    import queue as _queue
    import threading

    if loop is None:
        loop = asyncio.get_event_loop()
    if executor is None:
        executor = _executor

    q: "_queue.Queue" = _queue.Queue(maxsize=256)
    done_holder: Dict[str, Any] = {"value": None, "error": None}

    def progress_cb(pct, stage=""):
        try:
            q.put_nowait(("progress", float(pct or 0), str(stage or "")))
        except _queue.Full:
            pass

    def run():
        try:
            r = gen_fn(progress_cb)
            done_holder["value"] = r
        except Exception as e:
            done_holder["error"] = e
        finally:
            try:
                q.put_nowait(("__done__", None, None))
            except Exception:
                pass

    async def event_stream():
        # 在后台线程跑同步任务
        fut = loop.run_in_executor(executor, run)
        try:
            while True:
                try:
                    item = await loop.run_in_executor(None, q.get, True, 0.2)
                except _queue.Empty:
                    if fut.done():
                        # 排空剩余事件
                        break
                    continue
                if not item:
                    continue
                if item[0] == "__done__":
                    break
                _, pct, stage = item
                yield _sse_format("progress", {"pct": pct, "stage": stage})

            # 等待后台任务真正结束 (避免 race)
            try:
                await asyncio.wait_for(fut, timeout=timeout)
            except asyncio.TimeoutError:
                yield _sse_format("error", {"message": f"规划超时（{timeout}s）"})
                return

            if done_holder["error"] is not None:
                e = done_holder["error"]
                yield _sse_format("error", {"message": f"{type(e).__name__}: {e}"})
                return

            # 最终结果由调用方决定如何解析; 这里给一个哨兵
            yield _sse_format("result", done_holder["value"])
        finally:
            if not fut.done():
                fut.cancel()

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


def _pci_quality_after_plan(
    conflicts: List[Dict[str, Any]],
    *,
    check_mod30: bool = True,
    directional_filter: bool = True,
    ecgi_filter: Optional[Set[str]] = None,
) -> Dict[str, Any]:
    return build_pci_quality_report(
        STATE.cells,
        conflicts,
        ecgi_filter=ecgi_filter,
        check_mod30=check_mod30,
        directional_filter=directional_filter,
    )


def _pci_quality_for_planned_ecgis(
    planned_ecgis: List[str],
    *,
    check_mod30: bool = True,
    directional_filter: bool = True,
) -> Dict[str, Any]:
    """单站/批量新建小区：仅对 planned_ecgis 生成质量报告（邻区参考全网 STATE）。"""
    ecgis = {e for e in planned_ecgis if e}
    merged = _merge_runtime_cells()
    conflicts = collect_conflicts(merged, use_original_pci=False,
                                 directional_filter=directional_filter)
    return build_pci_quality_report(
        merged,
        conflicts,
        ecgi_filter=ecgis if ecgis else None,
        check_mod30=check_mod30,
        directional_filter=directional_filter,
    )


def _serialize_planned_cells(
    cells: List[Dict[str, Any]],
    quality_by_ecgi: Optional[Dict[str, Dict[str, Any]]] = None,
) -> List[Dict[str, Any]]:
    out = []
    for c in cells:
        item = {
            "ecgi": c.get("ecgi"),
            "name": c.get("name"),
            "site_name": c.get("site_name"),
            "phy_name": c.get("phy_name"),
            "ant_name": c.get("ant_name"),
            "manufacturer": c.get("manufacturer"),
            "oms_name": c.get("oms_name"),
            "rat": c.get("rat"),
            "freq_band": c.get("freq_band") or c.get("freq_band_label"),
            "lat": c.get("lat"),
            "lon": c.get("lon"),
            "azimuth": c.get("azimuth"),
            "beamwidth": c.get("beamwidth", c.get("beam")),
            "beam": c.get("beam"),
            "radius": c.get("radius"),
            "coverage_radius": c.get("coverage_radius"),
            "pci": c.get("pci"),
            "new_pci": c.get("new_pci"),
            "plan_site_type": c.get("plan_site_type"),
            "site_type_label": c.get("site_type_label"),
            "n_sectors": c.get("n_sectors", 1),
            "sector_index": c.get("sector_index"),
            "base_azimuth": c.get("base_azimuth"),
            "tac": c.get("tac"),
            "earfcn": c.get("earfcn"),
            "cell_id": c.get("cell_id"),
            "bandwidth": c.get("bandwidth"),
            "locked": c.get("locked", False),
            "pci_candidates": c.get("pci_candidates", []),
            "pci_candidates_primary": c.get("pci_candidates_primary"),
            "pci_groups": c.get("pci_groups", []),
        }
        if quality_by_ecgi and c.get("ecgi") in quality_by_ecgi:
            item["pci_quality"] = quality_by_ecgi[c["ecgi"]]
        elif c.get("pci_quality"):
            item["pci_quality"] = c.get("pci_quality")
        out.append(item)
    return out


def _resolve_single_site_plan_kwargs(req: SingleSitePlanRequest) -> Dict[str, Any]:
    """请求体未传的项用「规划默认参数」页持久化配置"""
    pci_def = pci_defaults_dict()
    score_thr = req.score_threshold
    if score_thr is None:
        score_thr = single_site_default_score_threshold()
    planning_mode = req.planning_mode or pci_def["planning_mode"]
    return {
        "lat": req.lat,
        "lon": req.lon,
        "rat": req.rat,
        "freq_band": req.freq_band,
        "plan_site_type": req.plan_site_type,
        "n_sectors": req.n_sectors,
        "base_azimuth": req.base_azimuth,
        "name_hint": req.name_hint,
        "site_name": req.site_name,
        "earfcn": req.earfcn,
        "tac": req.tac,
        "nbr_plan_types": req.nbr_plan_types,
        "engine": req.engine if req.engine else pci_def["engine"],
        "reuse_distance_km": req.reuse_distance_km,
        "check_mod6": req.check_mod6,
        "check_mod30": req.check_mod30,
        "use_beam_overlap_score": req.use_beam_overlap_score,
        "score_threshold": score_thr,
        "planning_mode": planning_mode,
        "directional_filter": req.directional_filter,
    }


def _purge_planned_temp_cells(state_cells: List[Dict[str, Any]]) -> int:
    """
    清理内存中的 PLAN-* 临时小区（_temp_cells；兼容旧版 STATE 内 is_temp）。
    返回被清理的临时小区数量.
    """
    before = len(_temp_cells)
    _temp_cells.clear()
    if state_cells is not STATE.cells:
        return before
    n_state = len(STATE.cells)
    STATE.cells[:] = [c for c in STATE.cells if not c.get("is_temp", False)]
    return before + (n_state - len(STATE.cells))


@app.post("/api/plan/single")
async def api_plan_single(req: SingleSitePlanRequest):
    """
    单站规划: 追加 N 个小区到 STATE.cells + PCI + 邻区;
    局部模式避免全量 O(N²) 重算，在 19K+ 小区下从分钟级降至秒级。
    """
    import asyncio

    _require_cells_loaded()
    if not try_acquire_plan_lock("plan_single"):
        return _plan_busy_response()

    _purge_planned_temp_cells(STATE.cells)
    kw = _resolve_single_site_plan_kwargs(req)

    loop = asyncio.get_event_loop()
    try:
        try:
            result = await asyncio.wait_for(
                loop.run_in_executor(
                    _executor,
                    lambda: plan_single_site(_merge_runtime_cells(), **kw),
                ),
                timeout=60.0,
            )
        except asyncio.TimeoutError:
            raise HTTPException(
                status_code=504,
                detail="单站规划超时（60s），请尝试减小 reuse_distance_km 或减少扇区数",
            )

        _temp_cells.extend(result["planned_cells"])

        pci_quality = _pci_quality_for_planned_ecgis(
            result["planned_ecgis"],
            check_mod30=req.check_mod30,
            directional_filter=req.directional_filter,
        )
        q_by_ecgi = {r["ecgi"]: r for r in pci_quality.get("cells", []) if r.get("ecgi")}
        planned = _serialize_planned_cells(result["planned_cells"], q_by_ecgi)

        if req.persist:
            try:
                db_save_all(STATE.cells)
            except Exception as e:
                result.setdefault("log", []).append(f"[warn] DB 持久化失败: {e}")

        resp = {
            "success": True,
            "center": result["center"],
            "planned_cells": planned,
            "planned_ecgis": result["planned_ecgis"],
            "nbr_by_type": result["nbr_by_type"],
            "nbr_counts": {k: len(v) for k, v in result["nbr_by_type"].items()},
            "engine": result["engine"],
            "log": result["log"][-30:],
            "stats": {
                "planned_count": len(planned),
                "pci_stats": result.get("pci_stats", {}),
                "nbr_stats": result.get("nbr_stats", {}),
                "pci_quality_summary": pci_quality.get("summary", {}),
            },
            "pci_quality": pci_quality,
        }
        release_planning_temp(trim_log=STATE.plan_log)
        return resp
    finally:
        release_plan_lock()


@app.post("/api/plan/single/stream")
async def api_plan_single_stream(req: SingleSitePlanRequest):
    """
    单站规划 (SSE 流式):
    - 推送 progress 事件 (pct, stage)
    - 最后推送 result 事件 (与 /api/plan/single 同 schema)
    """
    _require_cells_loaded()
    if not try_acquire_plan_lock("plan_single_stream"):
        return _plan_busy_response()
    _purge_planned_temp_cells(STATE.cells)

    def _run(progress_cb):
        try:
            kw = _resolve_single_site_plan_kwargs(req)
            kw["progress_cb"] = progress_cb
            result = plan_single_site(_merge_runtime_cells(), **kw)
            _temp_cells.extend(result["planned_cells"])
            pci_quality = _pci_quality_for_planned_ecgis(
                result["planned_ecgis"],
                check_mod30=req.check_mod30,
                directional_filter=req.directional_filter,
            )
            q_by_ecgi = {r["ecgi"]: r for r in pci_quality.get("cells", []) if r.get("ecgi")}
            planned = _serialize_planned_cells(result["planned_cells"], q_by_ecgi)
            if req.persist:
                try:
                    db_save_all(STATE.cells)
                except Exception as e:
                    result.setdefault("log", []).append(f"[warn] DB 持久化失败: {e}")
            release_planning_temp(trim_log=STATE.plan_log)
            return {
                "success": True,
                "center": result["center"],
                "planned_cells": planned,
                "planned_ecgis": result["planned_ecgis"],
                "nbr_by_type": result["nbr_by_type"],
                "nbr_counts": {k: len(v) for k, v in result["nbr_by_type"].items()},
                "engine": result["engine"],
                "log": result["log"][-30:],
                "stats": {
                    "planned_count": len(planned),
                    "pci_stats": result.get("pci_stats", {}),
                    "nbr_stats": result.get("nbr_stats", {}),
                    "pci_quality_summary": pci_quality.get("summary", {}),
                },
                "pci_quality": pci_quality,
            }
        finally:
            release_plan_lock()

    return _sse_stream_progress(_run, timeout=120.0)


@app.post("/api/plan/batch")
async def api_plan_batch(
    file: UploadFile = File(...),
    nbr_plan_types: str = Form(default="4G_4G,4G_5G,5G_4G,5G_5G"),
    engine: str = Form(default="legacy"),
    reuse_distance_km: float = Form(default=5.0),
    check_mod6: bool = Form(default=False),
    check_mod30: bool = Form(default=True),
    use_beam_overlap_score: bool = Form(default=False),
    directional_filter: bool = Form(default=True),
    auto_export: bool = Form(default=True),
    planning_mode: str = Form(default="pci+nbr"),
):
    """
    批量规划: 解析 xlsx + 局部规划 + 直接返回多 sheet xlsx
    """
    import asyncio

    _require_cells_loaded()
    if not try_acquire_plan_lock("plan_batch"):
        return _plan_busy_response()
    content = await file.read()
    npt_list = [s.strip() for s in nbr_plan_types.split(",") if s.strip()]

    try:
        result = await asyncio.wait_for(
            asyncio.get_event_loop().run_in_executor(
                _executor,
                lambda: plan_batch_sites(
                    STATE.cells,
                    file_bytes=content,
                    filename=file.filename or "upload.xlsx",
                    nbr_plan_types=npt_list,
                    engine=engine,
                    reuse_distance_km=reuse_distance_km,
                    check_mod6=check_mod6,
                    check_mod30=check_mod30,
                    use_beam_overlap_score=use_beam_overlap_score,
                    planning_mode=planning_mode,
                    directional_filter=directional_filter,
                ),
            ),
            timeout=120.0,
        )
    except asyncio.TimeoutError:
        raise HTTPException(
            status_code=504,
            detail="批量规划超时（120s），请减少上传行数或减小 reuse_distance_km",
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"批量规划失败: {e}")
    finally:
        release_plan_lock()
        del content
        gc.collect()

    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error", "批量规划失败"))

    export_cells = result.get("export_cells") or STATE.cells

    # 生成多 sheet xlsx（不写 STATE / 不入库）
    xlsx_blob = export_plan_split_sheets(
        export_cells, result["planned_ecgis"], nbr_plan_types=npt_list
    )

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    ascii_name = f"batch_plan_{ts}.xlsx"
    cn_name = f"批量规划结果_{ts}.xlsx"
    from urllib.parse import quote
    headers = {
        "Content-Disposition": (
            f'attachment; filename="{ascii_name}"; '
            f"filename*=UTF-8''{quote(cn_name)}"
        ),
    }
    # 同时回传 JSON 信息在 header 里 (供前端展示)
    import json as _json
    headers["X-Plan-Stats"] = _json.dumps({
        "planned": result["stats"]["planned"],
        "truncated": result["stats"]["truncated"],
        "invalid_rows": result["stats"]["invalid_rows"],
        "engine": result["engine"],
    })

    return Response(
        xlsx_blob,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.post("/api/plan/batch/stream")
async def api_plan_batch_stream(
    file: UploadFile = File(...),
    nbr_plan_types: str = Form(default="4G_4G,4G_5G,5G_4G,5G_5G"),
    engine: str = Form(default="legacy"),
    reuse_distance_km: float = Form(default=5.0),
    check_mod6: bool = Form(default=False),
    check_mod30: bool = Form(default=True),
    use_beam_overlap_score: bool = Form(default=False),
    directional_filter: bool = Form(default=True),
    planning_mode: str = Form(default="pci+nbr"),
):
    """
    批量规划 (SSE 流式):
    - 推送 progress 事件
    - 完成推送 done 事件 { session_id, filename, stats }
    - 前端随后用 GET /api/plan/batch/result/{session_id} 拉取 xlsx
    """
    import json as _json
    from urllib.parse import quote

    content = await file.read()
    npt_list = [s.strip() for s in nbr_plan_types.split(",") if s.strip()]
    session_id = uuid.uuid4().hex[:12]
    # 清理过期临时文件
    _cleanup_batch_results()

    def _run(progress_cb):
        result = plan_batch_sites(
            STATE.cells,
            file_bytes=content,
            filename=file.filename or "upload.xlsx",
            nbr_plan_types=npt_list,
            engine=engine,
            reuse_distance_km=reuse_distance_km,
            check_mod6=check_mod6,
            check_mod30=check_mod30,
            use_beam_overlap_score=use_beam_overlap_score,
            planning_mode=planning_mode,
            progress_cb=progress_cb,
            directional_filter=directional_filter,
        )
        if not result.get("success"):
            raise RuntimeError(result.get("error", "批量规划失败"))
        # 批量规划仅内存计算，不修改 STATE.cells、不写 cells 表
        result.setdefault("log", []).append(
            "[批量规划] 结果仅导出 xlsx，未写入工参库"
        )
        # 生成 xlsx
        progress_cb(98, "正在生成 xlsx…")
        export_cells = result.get("export_cells") or STATE.cells
        xlsx_blob = export_plan_split_sheets(
            export_cells, result["planned_ecgis"], nbr_plan_types=npt_list
        )
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        cn_name = f"批量规划结果_{ts}.xlsx"
        # 写入临时文件, 供后续下载
        out_path = BATCH_RESULT_DIR / f"{session_id}.bin"
        meta_path = BATCH_RESULT_DIR / f"{session_id}.json"
        out_path.write_bytes(xlsx_blob)
        meta_path.write_text(
            _json.dumps({"filename": cn_name, "stats": result.get("stats", {}), "engine": result.get("engine")}, ensure_ascii=False),
            encoding="utf-8",
        )
        progress_cb(100, "完成")
        return {
            "session_id": session_id,
            "filename": cn_name,
            "stats": result.get("stats", {}),
            "engine": result.get("engine"),
        }

    # 自定义 SSE 流 (yield progress + done)
    import asyncio
    import queue as _queue

    q: "_queue.Queue" = _queue.Queue(maxsize=256)
    done_holder: Dict[str, Any] = {"value": None, "error": None}

    def progress_cb(pct, stage=""):
        try: q.put_nowait(("progress", float(pct or 0), str(stage or "")))
        except _queue.Full: pass

    def run_bg():
        try:
            r = _run(progress_cb)
            done_holder["value"] = r
        except Exception as e:
            done_holder["error"] = e
        finally:
            try: q.put_nowait(("__done__", None, None))
            except Exception: pass

    loop = asyncio.get_event_loop()
    fut = loop.run_in_executor(_executor, run_bg)

    async def event_stream():
        try:
            while True:
                try:
                    item = await loop.run_in_executor(None, q.get, True, 0.2)
                except _queue.Empty:
                    if fut.done():
                        break
                    continue
                if not item: continue
                if item[0] == "__done__": break
                _, pct, stage = item
                yield _sse_format("progress", {"pct": pct, "stage": stage})
            try:
                await asyncio.wait_for(fut, timeout=240.0)
            except asyncio.TimeoutError:
                yield _sse_format("error", {"message": "批量规划超时（240s）"})
                return
            if done_holder["error"] is not None:
                e = done_holder["error"]
                yield _sse_format("error", {"message": f"{type(e).__name__}: {e}"})
                return
            yield _sse_format("done", done_holder["value"])
        finally:
            if not fut.done():
                fut.cancel()

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


@app.get("/api/plan/batch/result/{session_id}")
async def api_plan_batch_result(session_id: str):
    """
    拉取批量规划 SSE 完成后的 xlsx 文件 (一次有效, 下载完删除).
    """
    import json as _json
    from urllib.parse import quote
    bin_path = BATCH_RESULT_DIR / f"{session_id}.bin"
    meta_path = BATCH_RESULT_DIR / f"{session_id}.json"
    if not bin_path.exists():
        raise HTTPException(status_code=404, detail="结果已过期或不存在, 请重新规划")
    blob = bin_path.read_bytes()
    filename = "batch_plan.xlsx"
    if meta_path.exists():
        try:
            m = _json.loads(meta_path.read_text(encoding="utf-8"))
            filename = m.get("filename", filename)
        except Exception:
            pass
    # 删除临时文件 (一次性)
    try: bin_path.unlink()
    except Exception: pass
    try: meta_path.unlink()
    except Exception: pass
    ascii_name = re.sub(r"[^\x00-\x7f]+", "_", filename) or "batch_plan.xlsx"
    headers = {
        "Content-Disposition": (
            f'attachment; filename="{ascii_name}"; '
            f"filename*=UTF-8''{quote(filename)}"
        ),
    }
    return Response(
        blob,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.post("/api/export/split")
async def api_export_split(req: ExportSplitRequest):
    """
    导出规划结果 (多 sheet xlsx): PCI规划表 + 邻区-<类型>
    """
    try:
        xlsx_blob = export_plan_split_sheets(
            STATE.cells, req.planned_ecgis, nbr_plan_types=req.nbr_plan_types
        )
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        ascii_name = f"plan_split_{ts}.xlsx"
        cn_name = f"规划分sheet_{ts}.xlsx"
        from urllib.parse import quote
        headers = {
            "Content-Disposition": (
                f'attachment; filename="{ascii_name}"; '
                f"filename*=UTF-8''{quote(cn_name)}"
            ),
        }
        return Response(
            xlsx_blob,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {e}")


@app.post("/api/interference/analyze")
async def api_interference_analyze(req: InterferenceRequest):
    """
    扇区干扰分析 (同频/邻频/PCI 三检)
    - 可选 center_ecgi + radius_km: 仅分析此小区半径内
    - 可选 area:                前端圈选区域 (rect/circle)
    - 可选 rat / freq_band:     按 4G/5G 频段分桶过滤
    """
    try:
        cells_for_analysis = STATE.cells

        # 圈选区域过滤 (rect / circle / polygon)
        if req.area:
            from geo_utils import point_in_area
            area = req.area
            cells_for_analysis = [
                c for c in cells_for_analysis
                if point_in_area(c["lat"], c["lon"], area)
            ]
        elif req.center_ecgi and req.radius_km:
            from geo_utils import vincenty_distance
            target = next((c for c in STATE.cells if c.get("ecgi") == req.center_ecgi), None)
            if not target:
                raise HTTPException(status_code=404, detail=f"未找到 ECGI: {req.center_ecgi}")
            cells_for_analysis = [
                c for c in STATE.cells
                if vincenty_distance(target["lat"], target["lon"], c["lat"], c["lon"]) / 1000.0 <= req.radius_km
            ]

        result = analyze_interference(
            STATE.cells,
            interference_distance_km=req.interference_distance_km,
            overlap_threshold=req.overlap_threshold,
            detect_co_channel=req.detect_co_channel,
            detect_adjacent_channel=req.detect_adjacent_channel,
            detect_pci_collision=req.detect_pci_collision,
            detect_mod3=req.detect_mod3,
            detect_mod6=req.detect_mod6,
            rat_filter=req.rat,
            freq_band_filter=req.freq_band,
            area=req.area,
        )
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"干扰分析失败: {e}")


@app.post("/api/interference/export")
async def api_interference_export(req: InterferenceRequest):
    """
    导出干扰分析报告 (xlsx)
    """
    try:
        cells_for_analysis = STATE.cells

        # 圈选区域过滤 (与 /analyze 同步)
        if req.area:
            area = req.area
            atype = area.get("type")
            if atype == "rect" and area.get("lat1") is not None and area.get("lat2") is not None:
                la1, la2 = min(area["lat1"], area["lat2"]), max(area["lat1"], area["lat2"])
                lo1, lo2 = min(area["lon1"], area["lon2"]), max(area["lon1"], area["lon2"])
                cells_for_analysis = [
                    c for c in cells_for_analysis
                    if la1 <= c["lat"] <= la2 and lo1 <= c["lon"] <= lo2
                ]
            elif atype == "circle" and area.get("radius_km"):
                from geo_utils import vincenty_distance
                lat0, lon0 = area["lat"], area["lon"]
                r_km = float(area["radius_km"])
                cells_for_analysis = [
                    c for c in cells_for_analysis
                    if vincenty_distance(lat0, lon0, c["lat"], c["lon"]) / 1000.0 <= r_km
                ]
        elif req.center_ecgi and req.radius_km:
            from geo_utils import vincenty_distance
            target = next((c for c in STATE.cells if c.get("ecgi") == req.center_ecgi), None)
            if not target:
                raise HTTPException(status_code=404, detail=f"未找到 ECGI: {req.center_ecgi}")
            cells_for_analysis = [
                c for c in STATE.cells
                if vincenty_distance(target["lat"], target["lon"], c["lat"], c["lon"]) / 1000.0 <= req.radius_km
            ]

        result = analyze_interference(
            STATE.cells,
            interference_distance_km=req.interference_distance_km,
            overlap_threshold=req.overlap_threshold,
            detect_co_channel=req.detect_co_channel,
            detect_adjacent_channel=req.detect_adjacent_channel,
            detect_pci_collision=req.detect_pci_collision,
            detect_mod3=req.detect_mod3,
            detect_mod6=req.detect_mod6,
            rat_filter=req.rat,
            freq_band_filter=req.freq_band,
            area=req.area,
        )
        xlsx_blob = export_interference_report(result["issues"], result["mitigation"])
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        ascii_name = f"interference_{ts}.xlsx"
        cn_name = f"干扰分析报告_{ts}.xlsx"
        from urllib.parse import quote
        headers = {
            "Content-Disposition": (
                f'attachment; filename="{ascii_name}"; '
                f"filename*=UTF-8''{quote(cn_name)}"
            ),
        }
        return Response(
            xlsx_blob,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出干扰分析失败: {e}")


# =========================
# 配置导入 API
# =========================
class ColumnConfigItem(BaseModel):
    column_src: str
    column_dst: str
    data_type: str = "TEXT"
    is_pk: bool = False
    is_enabled: bool = True


class ColumnConfigRequest(BaseModel):
    sheet_name: str
    columns: List[ColumnConfigItem]
    enabled: bool = True
    description: str = ""
    unique_keys: Optional[List[str]] = None


@app.post("/api/config/upload")
async def api_config_upload(
    file: UploadFile = File(...),
    sheets: str = Form(default=""),
):
    """
    上传配置Excel文件，解析指定sheet并存入数据库

    - sheets: 逗号分隔的sheet名，为空则使用配置中的启用sheet
    """
    content = await file.read()
    if not content:
        raise HTTPException(400, "文件为空")

    # 解析sheet列表
    sheet_list = [s.strip() for s in sheets.split(",") if s.strip()] if sheets else None

    # 解析Excel
    result = parse_config_excel(content, file.filename or "upload.xlsx", sheet_list)
    if not result.get("success"):
        raise HTTPException(400, result.get("error", "解析失败"))

    # 存入数据库
    import_stats = {"sheets": [], "total_rows": 0, "errors": []}
    for sheet_name, sheet_data in result.get("sheets", {}).items():
        try:
            row_count = save_config_sheet(
                sheet_name,
                sheet_data["columns"],
                sheet_data["rows"],
                sheet_data["pk_columns"],
                file.filename or "",
            )
            import_stats["sheets"].append({
                "name": sheet_name,
                "row_count": row_count,
                "columns": len(sheet_data["columns"]),
            })
            import_stats["total_rows"] += row_count
        except Exception as e:
            import_stats["errors"].append(f"{sheet_name}: {e}")
            logger.error(f"保存配置sheet失败: {sheet_name}, {e}")

    return {
        "success": True,
        "filename": file.filename,
        "stats": result["stats"],
        "import_stats": import_stats,
    }


# ──────────────────────────────────────────────
# 进度事件结构（推送给前端）:
#   { "stage": "file_start",  "pct": 0,   "filename": "...", "file_idx": 0, "file_total": 1 }
#   { "stage": "parsing",     "pct": 5,   "filename": "...", "sheet": "X", "sheet_idx": 0, "sheet_total": 3 }
#   { "stage": "sheet_parsed","pct": ...,  "sheet": "X", "rows": 1234 }
#   { "stage": "saving",      "pct": ...,  "sheet": "X", "row": 100, "row_total": 1234 }
#   { "stage": "sheet_done",  "pct": ...,  "sheet": "X", "row_count": 1234 }
#   { "stage": "file_done",   "pct": 100, "filename": "...", "import_stats": {...} }
#   { "stage": "error",       "message": "..." }
# ──────────────────────────────────────────────

@app.post("/api/config/upload/stream")
async def api_config_upload_stream(
    file: UploadFile = File(...),
    sheets: str = Form(default=""),
):
    """
    上传并导入配置 Excel (SSE 流式)，推送分阶段进度:
      - 解析阶段 (0-15%): 解析 sheet 结构 + 行
      - 写入阶段 (15-100%): 逐 sheet 写入数据库，按行数加权进度
    """
    import asyncio
    import queue as _queue

    content = await file.read()
    if not content:
        raise HTTPException(400, "文件为空")

    sheet_list = [s.strip() for s in sheets.split(",") if s.strip()] if sheets else None
    filename = file.filename or "upload.xlsx"

    q: "_queue.Queue" = _queue.Queue(maxsize=512)
    done_holder: Dict[str, Any] = {"value": None, "error": None}

    def emit(payload: Dict[str, Any]) -> None:
        try:
            q.put_nowait(("progress", payload))
        except _queue.Full:
            pass

    def _run() -> Dict[str, Any]:
        try:
            # ── 阶段1: 解析 Excel（带 sheet 级别进度）──
            import openpyxl
            import io
            emit({"stage": "parsing", "pct": 2, "filename": filename, "status": "正在打开 Excel..."})

            # 自定义解析: 不使用 parse_config_excel 一次性解析，改为逐 sheet 解析
            # 这样可以发出 sheet 粒度的进度
            from config_imports import get_sheet_columns, is_sheet_enabled
            from config_parser import parse_single_sheet, apply_sheet_unique_keys, load_workbook_safe

            try:
                wb = load_workbook_safe(content, data_only=True)
            except Exception as e:
                raise RuntimeError(f"Excel文件读取失败: {e}")

            if sheet_list is None:
                # 用配置中启用的 sheet
                from config_imports import get_enabled_sheets
                sheet_list_local = get_enabled_sheets()
            else:
                sheet_list_local = sheet_list

            # 过滤存在的 sheet
            target_sheets = [s for s in sheet_list_local if s in wb.sheetnames]
            skipped = [s for s in sheet_list_local if s not in wb.sheetnames]

            if not target_sheets:
                wb.close()
                raise RuntimeError("没有任何需要导入的 sheet")

            # 阶段1: 0-15%  (解析阶段)
            # 阶段2: 15-100% (写入阶段, 按行数加权)
            parse_start_pct = 2
            parse_end_pct = 15
            save_start_pct = 15
            save_end_pct = 100

            parsed_sheets: Dict[str, Dict[str, Any]] = {}
            total_sheets = len(target_sheets)

            # ── 解析阶段 ──
            for i, sn in enumerate(target_sheets):
                emit({
                    "stage": "parsing",
                    "pct": parse_start_pct + (parse_end_pct - parse_start_pct) * i / max(total_sheets, 1),
                    "filename": filename,
                    "sheet": sn,
                    "sheet_idx": i,
                    "sheet_total": total_sheets,
                    "status": f"解析 sheet: {sn}",
                })

                # 检查列配置
                column_map = get_sheet_columns(sn)
                if not column_map:
                    emit({"stage": "warn", "filename": filename, "sheet": sn,
                          "message": f"sheet '{sn}' 无列配置，跳过"})
                    continue
                if not is_sheet_enabled(sn):
                    emit({"stage": "warn", "filename": filename, "sheet": sn,
                          "message": f"sheet '{sn}' 未启用，跳过"})
                    continue

                try:
                    ws = wb[sn]
                    sheet_data = parse_single_sheet(ws, column_map)
                    if sheet_data.get("error"):
                        emit({"stage": "warn", "filename": filename, "sheet": sn,
                              "message": f"解析失败: {sheet_data['error']}"})
                        continue
                    apply_sheet_unique_keys(sn, column_map, sheet_data)
                    for w in sheet_data.get("unique_key_warnings") or []:
                        emit({"stage": "warn", "filename": filename, "sheet": sn, "message": w})
                    parsed_sheets[sn] = sheet_data
                    emit({
                        "stage": "sheet_parsed",
                        "pct": parse_start_pct + (parse_end_pct - parse_start_pct) * (i + 1) / max(total_sheets, 1),
                        "filename": filename,
                        "sheet": sn,
                        "sheet_idx": i,
                        "sheet_total": total_sheets,
                        "rows": sheet_data["row_count"],
                        "columns": len(sheet_data["columns"]),
                    })
                except Exception as e:
                    emit({"stage": "warn", "filename": filename, "sheet": sn,
                          "message": f"解析异常: {e}"})

            wb.close()

            if not parsed_sheets:
                raise RuntimeError("所有 sheet 都解析失败")

            # ── 写入阶段: 按行数加权 ──
            total_rows_all = sum(sd["row_count"] for sd in parsed_sheets.values())
            if total_rows_all == 0:
                total_rows_all = 1  # 避免除零

            import_stats: Dict[str, Any] = {"sheets": [], "total_rows": 0, "errors": [], "skipped": skipped}

            # 已完成行数累计
            done_rows_cum = 0
            total_sheets_to_save = len(parsed_sheets)
            sheets_done = 0

            def _file_pct(extra: int = 0) -> float:
                """根据已完成行数计算当前文件的整体进度 (15-100%)"""
                cur = (done_rows_cum + extra) / total_rows_all
                cur = max(0.0, min(1.0, cur))
                return save_start_pct + (save_end_pct - save_start_pct) * cur

            for sn, sheet_data in parsed_sheets.items():
                sheet_total = sheet_data["row_count"]
                emit({
                    "stage": "saving",
                    "pct": _file_pct(0),
                    "filename": filename,
                    "sheet": sn,
                    "sheet_idx": sheets_done,
                    "sheet_total": total_sheets_to_save,
                    "row": 0,
                    "row_total": sheet_total,
                    "status": f"开始写入: {sn}",
                })

                # 写入时也推进行进度（每 ~1% 或每 500 行推一次）
                def _make_row_cb(sheet_name=sn, total=sheet_total):
                    last_emitted_row = [-1]
                    def cb(done_rows: int):
                        if total <= 0:
                            return
                        # 节流: 避免事件太密集
                        if done_rows != total and done_rows - last_emitted_row[0] < 500:
                            return
                        last_emitted_row[0] = done_rows
                        emit({
                            "stage": "saving",
                            "pct": _file_pct(done_rows),
                            "filename": filename,
                            "sheet": sheet_name,
                            "sheet_idx": sheets_done,
                            "sheet_total": total_sheets_to_save,
                            "row": done_rows,
                            "row_total": total,
                        })
                    return cb

                row_cb = _make_row_cb()

                try:
                    row_count = save_config_sheet(
                        sn,
                        sheet_data["columns"],
                        sheet_data["rows"],
                        sheet_data["pk_columns"],
                        filename,
                        progress_cb=row_cb,
                    )
                    # 写入结束, 推一次 100% 行进度
                    row_cb(sheet_total)

                    import_stats["sheets"].append({
                        "name": sn,
                        "row_count": row_count,
                        "columns": len(sheet_data["columns"]),
                    })
                    import_stats["total_rows"] += row_count
                    done_rows_cum += row_count
                    sheets_done += 1

                    emit({
                        "stage": "sheet_done",
                        "pct": _file_pct(0),
                        "filename": filename,
                        "sheet": sn,
                        "sheet_idx": sheets_done - 1,
                        "sheet_total": total_sheets_to_save,
                        "row_count": row_count,
                    })
                except Exception as e:
                    import_stats["errors"].append(f"{sn}: {e}")
                    logger.error(f"保存配置sheet失败: {sn}, {e}")
                    emit({"stage": "warn", "filename": filename, "sheet": sn,
                          "message": f"保存失败: {e}"})
                    # 失败也要推进 sheets_done 避免死循环
                    done_rows_cum += sheet_total
                    sheets_done += 1

            emit({
                "stage": "file_done",
                "pct": 100,
                "filename": filename,
                "import_stats": import_stats,
            })

            return {
                "success": True,
                "filename": filename,
                "import_stats": import_stats,
            }

        except Exception as e:
            logger.error(f"导入失败: {e}", exc_info=True)
            emit({"stage": "error", "message": str(e), "filename": filename})
            raise

    def run_bg():
        try:
            r = _run()
            done_holder["value"] = r
        except Exception as e:
            done_holder["error"] = e
        finally:
            try:
                q.put_nowait(("__done__", None))
            except Exception:
                pass

    loop = asyncio.get_event_loop()
    fut = loop.run_in_executor(_executor, run_bg)

    async def event_stream():
        try:
            while True:
                try:
                    item = await loop.run_in_executor(None, q.get, True, 0.2)
                except _queue.Empty:
                    if fut.done():
                        break
                    continue
                if not item:
                    continue
                if item[0] == "__done__":
                    break
                _, payload = item
                yield _sse_format("progress", payload)
            # 等待后台任务完成
            try:
                await asyncio.wait_for(fut, timeout=600.0)
            except asyncio.TimeoutError:
                yield _sse_format("error", {"message": "导入超时（600s）"})
                return
            if done_holder["error"] is not None:
                e = done_holder["error"]
                yield _sse_format("error", {"message": f"{type(e).__name__}: {e}"})
                return
            yield _sse_format("done", done_holder["value"])
        finally:
            if not fut.done():
                fut.cancel()

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


@app.get("/api/config/preview")
async def api_config_preview(filename: str = Query(...)):
    """
    预览Excel文件的sheet列表（不导入）

    - filename: 网管文件目录中的文件名
    """
    config_dir = ROOT_DIR / "网管文件"
    file_path = config_dir / filename

    if not file_path.exists():
        raise HTTPException(404, f"文件不存在: {filename}")

    try:
        content = file_path.read_bytes()
        result = list_excel_sheets(content, filename)
        return result
    except Exception as e:
        raise HTTPException(500, f"读取文件失败: {e}")


@app.post("/api/config/parse-excel")
async def api_config_parse_excel(file: UploadFile = File(...)):
    """
    上传并解析Excel文件，返回所有sheet的列名信息（用于配置）

    优化: 大文件 (>>100MB) 跳过行数统计，仅返回 sheet 列表 + 列名，
    避免 openpyxl 解析完整工作簿时长时间占用 CPU。
    """
    import io
    import openpyxl
    import time

    t0 = time.time()
    try:
        content = await file.read()
        file_size = len(content)
        logger.info(f"开始解析 Excel: {file.filename} ({file_size / 1024 / 1024:.1f} MB)")

        # 仅解析 sheet 元数据（不读任何行）
        # keep_links=False 减少内存；data_only=True 取缓存值
        from config_parser import load_workbook_safe
        wb = load_workbook_safe(content, data_only=True, keep_links=False)
        sheets = []

        # 大文件阈值：>50MB 跳过行数统计（仅返回列名）
        skip_row_count = file_size > 50 * 1024 * 1024

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            columns = []
            row_count = 0

            try:
                first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if first_row:
                    columns = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(first_row)]
            except Exception:
                columns = []

            if not skip_row_count:
                for _ in ws.iter_rows(min_row=2, values_only=True):
                    row_count += 1
                    if row_count > 10000:  # 限制计数
                        break

            sheets.append({
                "name": sheet_name,
                "columns": columns,
                "row_count": row_count if not skip_row_count else -1,
                "row_count_skipped": skip_row_count,
                "enabled": False,
            })

        wb.close()

        # 对比现有配置，标记已启用的sheet
        enabled_sheets = set(get_enabled_sheets())
        for s in sheets:
            if s["name"] in enabled_sheets:
                s["enabled"] = True

        logger.info(f"解析完成: {file.filename} 共 {len(sheets)} 个 sheet, 耗时 {time.time() - t0:.2f}s")
        return {"success": True, "sheets": sheets, "filename": file.filename}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"解析Excel失败: {e}", exc_info=True)
        raise HTTPException(500, f"解析Excel失败: {e}")


@app.get("/api/config/tables")
async def api_config_tables():
    """列出所有已导入的配置表"""
    tables = list_config_tables()
    # 添加表的显示名（去掉cfg_前缀）
    for t in tables:
        t["display_name"] = t["table_name"].replace("cfg_", "")
    return {"success": True, "tables": tables}


@app.get("/api/config/data/{table_name}")
async def api_config_data(
    table_name: str,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=2000),
    keyword: str = Query(default=None),
):
    """查询配置表数据（分页）"""
    # table_name不含cfg_前缀
    result = get_config_table_data(table_name, page, page_size, keyword)
    return {"success": True, **result}


@app.get("/api/config/history")
async def api_config_history(limit: int = Query(default=100, ge=1, le=500)):
    """获取导入历史"""
    history = get_import_history(limit)
    return {"success": True, "history": history}


@app.delete("/api/config/table/{table_name}")
async def api_config_delete_table(table_name: str):
    """删除配置表"""
    success = delete_config_table(table_name)
    if not success:
        raise HTTPException(500, "删除失败")
    return {"success": True, "message": f"已删除表 cfg_{table_name}"}


class ConfigBatchDeleteTablesRequest(BaseModel):
    table_names: List[str]


@app.post("/api/config/tables/batch-delete")
async def api_config_batch_delete_tables(body: ConfigBatchDeleteTablesRequest):
    """批量删除配置表（table_names 为不含 cfg_ 前缀的显示名）"""
    names = [n.strip() for n in (body.table_names or []) if n and str(n).strip()]
    if not names:
        raise HTTPException(400, "请至少选择一个表")
    deleted: List[str] = []
    failed: List[Dict[str, str]] = []
    for name in names:
        if delete_config_table(name):
            deleted.append(name)
        else:
            failed.append({"table_name": name, "error": "删除失败"})
    return {
        "success": len(failed) == 0,
        "deleted": deleted,
        "failed": failed,
        "message": f"已删除 {len(deleted)} 个表" + (f"，{len(failed)} 个失败" if failed else ""),
    }


# ── 列配置管理（Web界面） ────────────────────
@app.get("/api/config/sheet-configs")
async def api_sheet_configs():
    """列出所有sheet的列配置概览"""
    # 合合数据库配置和YAML配置
    db_configs = list_sheet_configs()
    yaml_configs = list_all_config_sheets()

    # 以YAML配置为主，补充数据库信息
    result = []
    for yc in yaml_configs:
        dc = next((d for d in db_configs if d["sheet_name"] == yc["name"]), None)
        result.append({
            "name": yc["name"],
            "enabled": yc["enabled"],
            "description": yc["description"],
            "yaml_column_count": yc["column_count"],
            "unique_keys": yc.get("unique_keys") or [],
            "db_column_count": dc.get("column_count", 0) if dc else 0,
            "last_updated": dc.get("last_updated") if dc else None,
        })

    return {"success": True, "configs": result}


@app.get("/api/config/sheet-config/{sheet_name}")
async def api_get_sheet_config(sheet_name: str):
    """获取指定sheet的列配置详情（仅从 config.yaml 读取）"""
    from config_imports import get_sheet_unique_keys

    sheet_yaml = load_import_config().get("sheets", {}).get(sheet_name, {})
    yaml_cols = sheet_yaml.get("columns", {})
    unique_keys = get_sheet_unique_keys(sheet_name)

    if yaml_cols:
        # YAML 是唯一来源
        pk_set = set(unique_keys)
        cols = []
        for i, (src, dst) in enumerate(yaml_cols.items()):
            cols.append({
                "column_src": src,
                "column_dst": dst,
                "data_type": "TEXT",
                "is_pk": dst in pk_set,
                "is_enabled": True,
                "display_order": i,
            })
        return {
            "success": True,
            "source": "yaml",
            "columns": cols,
            "unique_keys": unique_keys,
        }
    else:
        return {"success": True, "source": "none", "columns": [], "unique_keys": []}


@app.post("/api/config/sheet-config")
async def api_save_sheet_config(req: ColumnConfigRequest):
    """保存sheet的列配置（Web界面管理）
    同步保存到数据库和YAML文件
    """
    try:
        cols_dict = [c.dict() for c in req.columns]
        # 保存到数据库
        db_count = save_column_config(req.sheet_name, cols_dict)
        # 同步保存到YAML文件
        yaml_ok = save_sheet_config_to_yaml(
            req.sheet_name,
            cols_dict,
            description=req.description,
            enabled=req.enabled,
            unique_keys=req.unique_keys,
        )
        return {"success": True, "saved_count": db_count, "yaml_saved": yaml_ok}
    except Exception as e:
        raise HTTPException(500, f"保存失败: {e}")


@app.delete("/api/config/sheet-config/{sheet_name}")
async def api_delete_sheet_config(sheet_name: str):
    """删除sheet的列配置（恢复为YAML默认）"""
    success = delete_sheet_config(sheet_name)
    return {"success": success}


@app.post("/api/config/reload")
async def api_config_reload():
    """重新加载YAML配置"""
    reload_config()
    return {"success": True, "message": "配置已重新加载"}


@app.post("/api/config/sync-cells")
async def api_config_sync_cells():
    """同步网管基础数据到 cells 表

    读取 cfg_CUEUtranCellFDDLTE / cfg_CUEUtranCellTDDLTE / cfg_EUtranCellFDD,
    提取 enbid 并拼接 cgi, 按映射表更新 cells 表的 pci/tac/earfcn/freq_band_ind.
  成功后从数据库重载 STATE.cells，工参页 /api/cells 立即可见「网管同步」状态。
    """
    try:
        result = sync_cells_from_config()
        if result.get("success"):
            cells, meta = db_load_all()
            STATE.cells = cells or []
            if meta.get("stats"):
                STATE.plan_stats = {**(STATE.plan_stats or {}), **meta["stats"]}
            _refresh_workparam_stats()
            synced_n = sum(1 for c in STATE.cells if c.get("pci_synced_at"))
            result["cells_in_memory"] = len(STATE.cells)
            result["workparam_synced_count"] = synced_n
            result["workparam_unsynced_count"] = len(STATE.cells) - synced_n
        return result
    except Exception as e:
        logger.error("同步网管基础数据失败: %s", e, exc_info=True)
        raise HTTPException(500, f"同步失败: {e}")


# =========================
# 静态前端挂载
# =========================
if config.frontend_dir.exists():
    app.mount("/static", StaticFiles(directory=str(config.frontend_dir)), name="static")

    @app.get("/", response_class=HTMLResponse)
    async def index():
        idx = config.frontend_dir / "index.html"
        if idx.exists():
            return HTMLResponse(idx.read_text(encoding="utf-8"))
        return HTMLResponse("<h1>网优百宝箱 v1.2.1</h1><p>前端页面缺失</p>")

    @app.get("/plan", response_class=HTMLResponse)
    async def plan_page():
        idx = config.frontend_dir / "plan.html"
        if idx.exists():
            return HTMLResponse(idx.read_text(encoding="utf-8"))
        return HTMLResponse("<h1>页面不存在</h1>")

    @app.get("/pci", response_class=HTMLResponse)
    async def pci_page():
        idx = config.frontend_dir / "pci.html"
        if idx.exists():
            return HTMLResponse(idx.read_text(encoding="utf-8"))
        return HTMLResponse("<h1>页面不存在</h1>")

    @app.get("/config", response_class=HTMLResponse)
    async def config_page():
        idx = config.frontend_dir / "config.html"
        if idx.exists():
            return HTMLResponse(idx.read_text(encoding="utf-8"))
        return HTMLResponse("<h1>页面不存在</h1>")

    @app.get("/workparams", response_class=HTMLResponse)
    async def workparams_page():
        idx = config.frontend_dir / "workparams.html"
        if idx.exists():
            return HTMLResponse(idx.read_text(encoding="utf-8"))
        return HTMLResponse("<h1>页面不存在</h1>")

    @app.get("/settings", response_class=HTMLResponse)
    async def settings_page():
        idx = config.frontend_dir / "settings.html"
        if idx.exists():
            return HTMLResponse(idx.read_text(encoding="utf-8"))
        return HTMLResponse("<h1>页面不存在</h1>")

    @app.get("/config-data", response_class=HTMLResponse)
    async def config_data_page():
        idx = config.frontend_dir / "config-data.html"
        if idx.exists():
            return HTMLResponse(idx.read_text(encoding="utf-8"))
        return HTMLResponse("<h1>页面不存在</h1>")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        app,
        host=config.server_host,
        port=config.server_port,
    )