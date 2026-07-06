"""
工参数据导入与清洗模块
支持Excel(.xlsx/.xls)与CSV,标准化字段,异常校验,返回有效小区列表与异常明细
"""
from __future__ import annotations

import io
import math
import re
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd

from sector_params import enrich_cell_with_sector, normalize_freq_band
from site_type_ext import to_plan_site_type

# 字段映射:用户表头 -> 内部字段
FIELD_ALIASES: Dict[str, str] = {
    "ECGI": "ecgi",
    "小区ID": "ecgi",
    "小区标识": "ecgi",
    "CGI": "ecgi",
    "NCGI": "ecgi",
    "小区名称": "name",
    "小区名": "name",
    "名称": "name",
    "制式": "rat",
    "网络制式": "rat",
    "系统": "rat",
    "zhishi": "rat",
    "ZhiShi": "rat",
    "频点": "earfcn",
    "频段号": "earfcn",
    "ARFCN": "earfcn",
    "NARFCN": "earfcn",
    "中心频率": "earfcn",
    "频段": "earfcn",
    # 详细频段字段 (5G: 使用频段/详细使用频段/关联频段; 4G: 详细使用频段)
    "详细使用频段": "freq_band_raw",
    "使用频段": "freq_band_raw",
    "关联频段": "freq_band_raw",
    "详细频段": "freq_band_raw",
    "频段标签": "freq_band_raw",
    "规划频段": "plan_freq_band",
    "频段(规划)": "plan_freq_band",
    "经度": "lon",
    "LON": "lon",
    "LONGITUDE": "lon",
    "lng": "lon",
    "纬度": "lat",
    "LAT": "lat",
    "LATITUDE": "lat",
    "方位角": "azimuth",
    "方向角": "azimuth",
    "波束方向": "azimuth",
    "覆盖半径": "radius",
    "半径": "radius",
    "站间距": "radius",
    "TAC": "tac",
    "跟踪区": "tac",
    "TA": "tac",
    "跟踪区码TAC": "tac",
    "PCI": "pci",
    "物理小区ID": "pci",
    "小区PCI": "pci",
    "物理小区识别码": "pci",
    "站点类型": "site_type",
    "站型": "site_type",
    "场景": "site_type",
    "覆盖类型": "site_type",
    "Station": "site_type",
    # site_name 不再从工参列头解析, 由 phy_name / name 派生
    # 保留内部字段用于 PCI/邻区规划分组
    # 物理站 (4G = 所属站点名称, 5G = 所属局站, 以及通用叫法)
    "物理站": "phy_name",
    "所属物理站": "phy_name",
    "物理站点": "phy_name",
    "所属站点名称": "phy_name",
    "所属站点": "phy_name",
    "所属局站": "phy_name",
    "局站": "phy_name",
    "局站名称": "phy_name",
    "所属基站名称": "phy_name",
    "所属基站": "phy_name",
    # 天线名称
    "天线名称": "ant_name",
    "天线": "ant_name",
    "天线名": "ant_name",
    "antenna": "ant_name",
    "Antenna": "ant_name",
    # 厂家
    "厂家": "manufacturer",
    "设备厂家": "manufacturer",
    "厂商": "manufacturer",
    "设备厂商": "manufacturer",
    "vendor": "manufacturer",
    "Vendor": "manufacturer",
    "VENDOR": "manufacturer",
    # 归属网管 (oms_name)
    "归属网管": "oms_name",
    "网管": "oms_name",
    "网管系统": "oms_name",
    "归属OMC": "oms_name",
    "OMC": "oms_name",
    "归属网管名称": "oms_name",
    # 单站/批量规划扩展字段
    "扇区数": "n_sectors",
    "sector_count": "n_sectors",
    "基方位角": "base_azimuth",
    "规划类型": "plan_site_type",
    "plan_type": "plan_site_type",
    "邻区规划": "nbr_plan_types",
    "邻区规划类型": "nbr_plan_types",
    "nbr_plan_types": "nbr_plan_types",
    "锁定": "locked",
    "locked": "locked",
    "邻区得分阈值": "nbr_score_threshold",
    "得分阈值": "nbr_score_threshold",
    "邻区阈值": "nbr_score_threshold",
    "score_threshold": "nbr_score_threshold",
}

# 内部标准字段
STD_FIELDS = [
    "ecgi", "name", "rat", "earfcn", "lon", "lat", "azimuth", "radius",
    "tac", "pci", "new_pci", "site_name", "phy_name", "ant_name", "manufacturer",
    "oms_name",
    "site_type", "freq_band_raw", "freq_band", "freq_band_label", "plan_freq_band",
    "bandwidth", "beam", "beamwidth", "pci_missing", "cell_id",
    "n_sectors", "plan_site_type", "base_azimuth", "locked", "nbr_score_threshold",
    "neighbors_json", "updated_at",
]

# 校验规则
PCI_RANGE = {"LTE": (0, 503), "NR": (0, 1007)}
RAT_ALIASES = {
    "LTE": "LTE", "4G": "LTE", "FDD-LTE": "LTE", "TDD-LTE": "LTE", "EUTRA": "LTE",
    "FDD": "LTE", "TDD": "LTE", "NB-IOT": "LTE", "NB-IoT": "LTE", "EUTRAN": "LTE",
    "NR": "NR", "5G": "NR", "5G NR": "NR", "gNB": "NR", "TDD-NR": "NR",
}
SITE_TYPE_ALIASES = {
    "陆地": "陆地", "室内": "室内", "近海": "近海", "海域": "近海",
    "LAND": "陆地", "INDOOR": "室内", "OFFSHORE": "近海", "SEA": "近海",
}


def _strip_template_header_label(header: str) -> str:
    """去掉下载模板表头后缀，如「小区名称[必填][开放]」→「小区名称」"""
    s = str(header).strip()
    s = re.sub(r"\[(必填|可选|枚举|开放)\]\s*$", "", s)
    s = re.sub(r"\[(必填|可选|枚举|开放)\]\s*$", "", s)
    return s.strip()


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """表头归一化:去除空格、括号、中文标点; 合并重复映射(保留首列)"""
    new_cols: Dict[str, str] = {}
    for c in df.columns:
        if c is None:
            continue
        cleaned = _strip_template_header_label(str(c).strip())
        if cleaned in FIELD_ALIASES:
            new_cols[c] = FIELD_ALIASES[cleaned]
            continue
        # 尝试去除括号/冒号/空格后匹配
        cleaned2 = re.sub(r"[\s()()\[\]【】:：]", "", cleaned)
        for k, v in FIELD_ALIASES.items():
            if re.sub(r"[\s()()\[\]【】:：]", "", k) == cleaned2:
                new_cols[c] = v
                break
    renamed = df.rename(columns=new_cols)
    # 出现重复目标列时: 合并为单列 (例如 4G=所属站点名称 + 5G=所属局站 都映射到 phy_name)
    # 按行取该列所有源中第一个非空值, 避免 groupby().first() 取到第一行的空值
    if renamed.columns.duplicated().any():
        # 收集每个目标列的源列名
        groups: Dict[str, List[str]] = {}
        for col in renamed.columns:
            groups.setdefault(col, []).append(col)
        # 用一个唯一前缀再 rename 的方法不如直接按列名分组合并 Series
        merged_data: Dict[str, pd.Series] = {}
        seen_order: List[str] = []
        for col, srcs in groups.items():
            if col not in merged_data:
                seen_order.append(col)
            if len(srcs) == 1:
                merged_data[col] = renamed[col].copy()
            else:
                # 每行取首个非空
                stack = renamed[srcs]   # DataFrame, 列 = srcs
                # 行内 first-non-null: 用 where + ffill
                # 先转置: (n_rows, n_srcs), 每行扫描
                def _pick_row(row: pd.Series) -> Any:
                    for v in row:
                        if v is not None and not (isinstance(v, float) and pd.isna(v)) and v != "":
                            return v
                    return row.iloc[0]
                merged_data[col] = stack.apply(_pick_row, axis=1)
        renamed = pd.DataFrame(merged_data, index=renamed.index)[seen_order]
    return renamed


def _coerce_value(field: str, value: Any) -> Tuple[Any, str]:
    """
    字段类型强转,返回(转换值,错误信息)
    """
    if value is None:
        return None, "为空"
    if isinstance(value, float) and math.isnan(value):
        return None, "为空"

    sval = str(value).strip()
    if field == "rat":
        v = RAT_ALIASES.get(sval.upper(), RAT_ALIASES.get(sval, None))
        return (v if v else None), ("" if v else f"未知制式: {sval}")
    if field == "site_type":
        v = SITE_TYPE_ALIASES.get(sval.upper(), SITE_TYPE_ALIASES.get(sval, "陆地"))
        return v, ""
    if field in ("lon", "lat", "azimuth", "radius", "earfcn", "tac", "pci"):
        try:
            num = float(sval) if field in ("lon", "lat", "azimuth", "radius") else int(float(sval))
            return num, ""
        except (ValueError, TypeError):
            return None, f"非数字: {sval}"
    if field == "n_sectors":
        try:
            n = int(float(sval))
            if n < 1 or n > 6:
                return None, f"扇区数越界(1-6): {n}"
            return n, ""
        except (ValueError, TypeError):
            return None, f"扇区数非数字: {sval}"
    if field == "base_azimuth":
        try:
            n = float(sval)
            if not (0 <= n < 360):
                return n % 360.0, ""
            return n, ""
        except (ValueError, TypeError):
            return None, f"基方位角非数字: {sval}"
    if field == "plan_site_type":
        # 自由字符串, 后续 to_plan_site_type 归一化
        return sval, ""
    if field == "nbr_plan_types":
        # 字符串以 | 或 , 分隔
        return sval, ""
    if field == "nbr_score_threshold":
        if not sval:
            return None, ""
        try:
            t = float(sval)
            if not (0.0 <= t <= 1.0):
                return None, f"邻区得分阈值越界(0-1): {t}"
            return round(t, 4), ""
        except (ValueError, TypeError):
            return None, f"邻区得分阈值非数字: {sval}"
    if field == "locked":
        if isinstance(value, bool):
            return value, ""
        if sval.lower() in ("true", "1", "yes", "locked", "是", "y", "t"):
            return True, ""
        if sval.lower() in ("false", "0", "no", "unlocked", "否", "n", "f", ""):
            return False, ""
        return False, ""
    if field in ("bandwidth", "beamwidth"):
        try:
            n = float(sval)
            return n, ""
        except (ValueError, TypeError):
            return None, f"非数字: {sval}"
    # manufacturer / oms_name / site_name / name / ant_name / phy_name / cell_id / new_pci / freq_band / freq_band_label / beam / plan_site_type / neighbors_json: 自由字符串
    return sval, ""


def _validate_cell(cell: Dict[str, Any]) -> List[str]:
    """校验单条小区数据,返回错误列表(已先做默认值填补)"""
    errors: List[str] = []
    if not cell.get("ecgi"):
        errors.append("ECGI为空")
    if not cell.get("name"):
        errors.append("小区名称为空")
    if cell.get("rat") not in ("LTE", "NR"):
        errors.append(f"制式异常: {cell.get('rat')}")
    lon, lat = cell.get("lon"), cell.get("lat")
    if lon is None or lat is None or not (-180 <= lon <= 180) or not (-90 <= lat <= 90):
        errors.append(f"经纬度异常: {lon},{lat}")
    az = cell.get("azimuth")
    if az is not None:
        # 容错: 360° 在通信工参里等同 0°, 大于 360 视为溢出模 360
        if az >= 360:
            cell["azimuth"] = az % 360.0
            az = cell["azimuth"]
        if not (0 <= az < 360):
            errors.append(f"方位角异常: {az}")
    # radius 可选: 工参常无此字段, 由频段/制式/站型映射填补
    r = cell.get("radius")
    if r is not None and (r <= 0 or r > 200000):
        errors.append(f"覆盖半径异常: {r}")
    pci = cell.get("pci")
    if pci is not None and pci >= 0:
        rng = PCI_RANGE.get(cell.get("rat"))
        if rng and not (rng[0] <= pci <= rng[1]):
            errors.append(f"PCI {pci} 超出{cell.get('rat')}范围{rng}")
    return errors


def parse_work_params(file_bytes: bytes, filename: str) -> Dict[str, Any]:
    """
    主入口:解析工参Excel/CSV
    返回:
      {
        "valid_cells": [...],
        "invalid_rows": [...],
        "stats": {"total": N, "valid": M, "invalid": K, "rat_counts": {...}}
      }
    """
    fname = (filename or "").lower()
    try:
        if fname.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(file_bytes), dtype=str, keep_default_na=False)
        elif fname.endswith(".xls"):
            df = pd.read_excel(io.BytesIO(file_bytes), dtype=str, keep_default_na=False, engine="xlrd")
        else:
            df = pd.read_excel(io.BytesIO(file_bytes), dtype=str, keep_default_na=False, engine="openpyxl")
    except Exception as e:
        return {
            "valid_cells": [],
            "invalid_rows": [],
            "stats": {"total": 0, "valid": 0, "invalid": 0, "rat_counts": {}},
            "error": f"文件读取失败: {e}",
        }

    df = _normalize_columns(df)

    # 批量模板第 2 行为「取值/范围」说明，上传时跳过（勿当数据行）
    if len(df) > 0:
        first = df.iloc[0]
        ecgi_val = str(first.get("ecgi", "") or "").strip()
        rat_val = str(first.get("rat", "") or "").strip()
        if ecgi_val.startswith("MCC-MNC") or (
            rat_val
            and rat_val not in ("LTE", "NR", "4G", "5G")
            and not any(ch.isdigit() for ch in rat_val[:3])
        ):
            df = df.iloc[1:].reset_index(drop=True)

    # 必填列: pci 可选（批量新建留空/-1，由 PCI 规划写入 new_pci 后导出）
    missing = [f for f in ("ecgi", "name", "rat", "lon", "lat", "azimuth") if f not in df.columns]
    if missing:
        return {
            "valid_cells": [],
            "invalid_rows": [{"row": 0, "errors": [f"缺少必填列: {','.join(missing)}"]}],
            "stats": {"total": 0, "valid": 0, "invalid": 0, "rat_counts": {}},
            "error": "缺少必填列",
        }

    valid: List[Dict[str, Any]] = []
    invalid: List[Dict[str, Any]] = []
    rat_counts: Dict[str, int] = {}

    for idx, row in df.iterrows():
        cell: Dict[str, Any] = {}
        for f in STD_FIELDS:
            if f in df.columns:
                v, _ = _coerce_value(f, row.get(f))
                cell[f] = v
            else:
                # 默认值填补
                if f == "site_type":
                    cell[f] = "陆地"
                elif f == "site_name":
                    cell[f] = ""  # 派生字段, 在 enrich_cell_with_sector 之后统一处理
                elif f == "phy_name":
                    cell[f] = ""
                elif f == "ant_name":
                    cell[f] = ""
                elif f == "manufacturer":
                    cell[f] = ""
                elif f == "oms_name":
                    cell[f] = ""
                elif f == "n_sectors":
                    cell[f] = 1
                elif f == "base_azimuth":
                    cell[f] = 0.0
                elif f == "plan_site_type":
                    cell[f] = None
                elif f == "nbr_plan_types":
                    cell[f] = None
                elif f == "nbr_score_threshold":
                    cell[f] = None
                elif f == "locked":
                    cell[f] = False
                elif f == "pci_missing":
                    cell[f] = False
                elif f == "updated_at":
                    cell[f] = None
                elif f == "neighbors_json":
                    cell[f] = "[]"
                else:
                    cell[f] = None

        # 容错填补: 方位角缺失默认 0, PCI 缺失记为 -1 标记
        if cell.get("azimuth") is None:
            cell["azimuth"] = 0.0
        if cell.get("pci") is None:
            cell["pci"] = -1
            cell["pci_missing"] = True

        # nbr_plan_types 拆分为列表
        raw_npt = cell.get("nbr_plan_types")
        if raw_npt:
            npt_list = [s.strip() for s in re.split(r"[|,]", str(raw_npt)) if s.strip()]
            cell["nbr_plan_types"] = npt_list
        else:
            cell["nbr_plan_types"] = None

        # plan_site_type 归一化
        cell["plan_site_type"] = to_plan_site_type(cell.get("plan_site_type") or cell.get("site_type"))

        errs = _validate_cell(cell)
        if errs:
            invalid.append({
                "row": int(idx) + 2,  # +2: 1-indexed + header row
                "raw": {k: (str(row.get(k)) if k in df.columns else "") for k in df.columns},
                "errors": errs,
            })
            continue

        # 派生字段
        # 频段标准化: 优先使用频段标签, 否则从频段号 + earfcn 推断
        freq_raw = cell.get("freq_band_raw")
        freq_band = normalize_freq_band(cell["rat"], freq_raw)
        if not freq_band and cell.get("earfcn") is not None:
            freq_band = _guess_freq_band(cell["rat"], cell.get("earfcn"))
        cell["freq_band"] = freq_band or "默认"

        # CASE 映射: 站点类型 + 制式 + 频段 → (beam, radius, 标签)
        enrich_cell_with_sector(cell)

        # site_name 派生: 优先级 1=phy_name, 优先级 2=name (按要求)
        # 工参里不再包含 site_name 列
        phy = cell.get("phy_name")
        if phy:
            cell["site_name"] = phy
        else:
            nm = cell.get("name") or ""
            # name 通常为 "XXX-1/2/3", 取前缀作为站点; 无 '-' 时原样使用
            cell["site_name"] = nm.rsplit("-", 1)[0] if "-" in nm else nm

        rat_counts[cell["rat"]] = rat_counts.get(cell["rat"], 0) + 1
        valid.append(cell)

    return {
        "valid_cells": valid,
        "invalid_rows": invalid,
        "stats": {
            "total": int(len(df)),
            "valid": len(valid),
            "invalid": len(invalid),
            "rat_counts": rat_counts,
        },
    }


def describe_file_rat_profile(
    filename: str,
    rat_counts: Dict[str, int],
    valid: int = 0,
) -> Dict[str, Any]:
    """
    根据解析结果与文件名推断工参文件的制式画像（4G/5G/混合）。
    """
    lte = int(rat_counts.get("LTE") or 0)
    nr = int(rat_counts.get("NR") or 0)
    fn = (filename or "").lower()

    if lte > 0 and nr > 0:
        kind = "mixed"
        label = "4G+5G 混合"
    elif nr > 0 and lte == 0:
        kind = "5G"
        label = "5G NR"
    elif lte > 0 and nr == 0:
        kind = "4G"
        label = "4G LTE"
    else:
        kind = "unknown"
        label = "未识别制式"
        if re.search(r"(?<![0-9])5g|nr|gnodeb|gnb", fn):
            kind, label = "5G", "5G（文件名推断）"
        elif re.search(r"(?<![0-9])4g|lte|enodeb|eutran", fn):
            kind, label = "4G", "4G（文件名推断）"

    return {
        "kind": kind,
        "label": label,
        "lte": lte,
        "nr": nr,
        "valid": valid,
    }


# 文件批量更新：仅 cgi + pci / tac / earfcn（空单元格表示不修改该字段）
BULK_UPDATE_FIELDS = ("ecgi", "pci", "tac", "earfcn")
_BULK_UPDATE_ALIASES: Dict[str, str] = {
    **{k: v for k, v in FIELD_ALIASES.items() if v in BULK_UPDATE_FIELDS},
    "cgi": "ecgi",
    "CGI": "ecgi",
    "ecgi": "ecgi",
    "ECGI": "ecgi",
    "pci": "pci",
    "PCI": "pci",
    "tac": "tac",
    "TAC": "tac",
    "earfcn": "earfcn",
    "EARFCN": "earfcn",
    "频点": "earfcn",
}


def normalize_ecgi_key(ecgi: Any) -> str:
    """工参关联键：去首尾空白，合并连续空白。"""
    if ecgi is None:
        return ""
    s = str(ecgi).strip()
    if not s:
        return ""
    return re.sub(r"\s+", "", s)


def _normalize_bulk_update_columns(df: pd.DataFrame) -> pd.DataFrame:
    new_cols: Dict[str, str] = {}
    for c in df.columns:
        if c is None:
            continue
        cleaned = _strip_template_header_label(str(c).strip())
        if cleaned in _BULK_UPDATE_ALIASES:
            new_cols[c] = _BULK_UPDATE_ALIASES[cleaned]
            continue
        cleaned2 = re.sub(r"[\s()()\[\]【】:：]", "", cleaned)
        for k, v in _BULK_UPDATE_ALIASES.items():
            if re.sub(r"[\s()()\[\]【】:：]", "", k) == cleaned2:
                new_cols[c] = v
                break
    renamed = df.rename(columns=new_cols)
    if renamed.columns.duplicated().any():
        groups: Dict[str, List[str]] = {}
        for col in renamed.columns:
            groups.setdefault(col, []).append(col)
        merged_data: Dict[str, pd.Series] = {}
        seen_order: List[str] = []
        for col, srcs in groups.items():
            if col not in merged_data:
                seen_order.append(col)
            if len(srcs) == 1:
                merged_data[col] = renamed[col].copy()
            else:
                stack = renamed[srcs]

                def _pick_row(row: pd.Series) -> Any:
                    for v in row:
                        if v is not None and not (isinstance(v, float) and pd.isna(v)) and str(v).strip() != "":
                            return v
                    return row.iloc[0]

                merged_data[col] = stack.apply(_pick_row, axis=1)
        renamed = pd.DataFrame(merged_data, index=renamed.index)[seen_order]
    return renamed


def _parse_optional_int_field(field: str, value: Any, rat: Optional[str]) -> Tuple[Optional[int], Optional[str]]:
    """解析批量更新中的可选整数字段；空值返回 (None, None) 表示跳过更新。"""
    if value is None:
        return None, None
    if isinstance(value, float) and math.isnan(value):
        return None, None
    sval = str(value).strip()
    if sval == "" or sval.lower() in ("nan", "null", "none", "-", "—"):
        return None, None
    try:
        num = int(float(sval))
    except (ValueError, TypeError):
        return None, f"{field} 非数字: {sval}"
    if field == "pci":
        if num < 0:
            return -1, None
        if rat in PCI_RANGE:
            lo, hi = PCI_RANGE[rat]
            if not (lo <= num <= hi):
                return None, f"PCI {num} 超出{rat}范围({lo}-{hi})"
        elif rat is None:
            if num > 1007:
                return None, f"PCI {num} 超出常见范围"
        return num, None
    if field == "tac":
        if num < 0 or num > 0xFFFF:
            return None, f"TAC 越界: {num}"
        return num, None
    if field == "earfcn":
        return num, None
    return num, None


def parse_bulk_sector_updates(file_bytes: bytes, filename: str) -> Dict[str, Any]:
    """
    解析「按 CGI 批量更新 PCI/TAC/EARFCN」文件 (csv/xlsx/xls)。
    返回:
      success, error?,
      rows: [{ecgi, pci?, tac?, earfcn?}],  # 仅包含文件中非空待更新字段
      invalid_rows, stats
    """
    fname = (filename or "").lower()
    try:
        if fname.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(file_bytes), dtype=str, keep_default_na=False)
        elif fname.endswith(".xls"):
            df = pd.read_excel(io.BytesIO(file_bytes), dtype=str, keep_default_na=False, engine="xlrd")
        else:
            df = pd.read_excel(io.BytesIO(file_bytes), dtype=str, keep_default_na=False, engine="openpyxl")
    except Exception as e:
        return {
            "success": False,
            "error": f"文件读取失败: {e}",
            "rows": [],
            "invalid_rows": [],
            "stats": {"total": 0, "valid": 0, "invalid": 0},
        }

    df = _normalize_bulk_update_columns(df)
    if "ecgi" not in df.columns:
        return {
            "success": False,
            "error": "缺少 CGI/ECGI 列（表头可为 cgi、CGI、ECGI、小区ID 等）",
            "rows": [],
            "invalid_rows": [],
            "stats": {"total": 0, "valid": 0, "invalid": 0},
        }

    if len(df) > 0:
        first = df.iloc[0]
        ecgi_val = str(first.get("ecgi", "") or "").strip()
        if ecgi_val.startswith("MCC-MNC") or ecgi_val.upper() in ("CGI", "ECGI", "示例"):
            df = df.iloc[1:].reset_index(drop=True)

    valid_rows: List[Dict[str, Any]] = []
    invalid: List[Dict[str, Any]] = []

    for idx, row in df.iterrows():
        ecgi = normalize_ecgi_key(row.get("ecgi"))
        if not ecgi:
            invalid.append({
                "row": int(idx) + 2,
                "ecgi": "",
                "errors": ["CGI/ECGI 为空"],
            })
            continue

        entry: Dict[str, Any] = {"ecgi": ecgi}
        row_errors: List[str] = []

        if "pci" in df.columns:
            pci, err = _parse_optional_int_field("pci", row.get("pci"), None)
            if err:
                row_errors.append(err)
            elif pci is not None:
                entry["pci"] = pci

        if "tac" in df.columns:
            tac, err = _parse_optional_int_field("tac", row.get("tac"), None)
            if err:
                row_errors.append(err)
            elif tac is not None:
                entry["tac"] = tac

        if "earfcn" in df.columns:
            earfcn, err = _parse_optional_int_field("earfcn", row.get("earfcn"), None)
            if err:
                row_errors.append(err)
            elif earfcn is not None:
                entry["earfcn"] = earfcn

        if row_errors:
            invalid.append({"row": int(idx) + 2, "ecgi": ecgi, "errors": row_errors})
            continue

        if len(entry) <= 1:
            invalid.append({
                "row": int(idx) + 2,
                "ecgi": ecgi,
                "errors": ["pci/tac/earfcn 至少填写一项"],
            })
            continue

        valid_rows.append(entry)

    return {
        "success": True,
        "rows": valid_rows,
        "invalid_rows": invalid,
        "stats": {
            "total": int(len(df)),
            "valid": len(valid_rows),
            "invalid": len(invalid),
        },
    }


def apply_bulk_sector_updates(
    cells: List[Dict[str, Any]],
    update_rows: List[Dict[str, Any]],
) -> Dict[str, Any]:
    """
    按 ECGI 关联更新内存工参中的 pci / tac / earfcn。
    返回 updated, not_found, skipped_duplicate_ecgi, updated_ecgis, not_found_ecgis
    """
    from datetime import datetime

    index: Dict[str, int] = {}
    for i, c in enumerate(cells):
        key = normalize_ecgi_key(c.get("ecgi"))
        if key and key not in index:
            index[key] = i

    updated = 0
    not_found: List[str] = []
    seen_file_ecgi: Dict[str, int] = {}
    skipped_duplicate = 0
    skipped_pci_invalid = 0
    updated_ecgis: List[str] = []
    ts = datetime.utcnow().isoformat(timespec="seconds")

    for row in update_rows:
        key = normalize_ecgi_key(row.get("ecgi"))
        if not key:
            continue
        if key in seen_file_ecgi:
            skipped_duplicate += 1
            continue
        seen_file_ecgi[key] = 1

        idx = index.get(key)
        if idx is None:
            not_found.append(key)
            continue

        cell = cells[idx]
        rat = cell.get("rat")
        changed = False

        if "pci" in row:
            pci = row["pci"]
            pci_ok = True
            if pci is not None and pci >= 0 and rat in PCI_RANGE:
                lo, hi = PCI_RANGE[rat]
                if not (lo <= pci <= hi):
                    pci_ok = False
                    skipped_pci_invalid += 1
            if pci_ok:
                cell["pci"] = pci
                if pci is not None and pci >= 0:
                    cell["pci_missing"] = False
                    cell["new_pci"] = pci
                elif pci == -1:
                    cell["pci_missing"] = True
                changed = True

        if "tac" in row:
            cell["tac"] = row["tac"]
            changed = True

        if "earfcn" in row:
            earfcn = row["earfcn"]
            cell["earfcn"] = earfcn
            if earfcn is not None and rat:
                guessed = _guess_freq_band(rat, earfcn)
                if guessed and guessed != "未知":
                    cell["freq_band"] = guessed
            changed = True

        if changed:
            cell["updated_at"] = ts
            cell.pop("pci_synced_at", None)
            updated += 1
            updated_ecgis.append(key)

    return {
        "updated": updated,
        "not_found": not_found,
        "not_found_count": len(not_found),
        "skipped_duplicate_ecgi": skipped_duplicate,
        "skipped_pci_invalid": skipped_pci_invalid,
        "updated_ecgis": updated_ecgis[:500],
        "not_found_ecgis": not_found[:500],
    }


def generate_bulk_sector_update_template() -> bytes:
    """生成 4 列更新模板 xlsx（含说明行）。"""
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "更新清单"
    headers = ["cgi", "pci", "tac", "earfcn"]
    ws.append(headers)
    ws.append([
        "MCC-MNC-eNB/gNB-CellId（与工参 ECGI 一致）",
        "物理小区ID，空表示不修改",
        "跟踪区码，空表示不修改",
        "频点/ARFCN，空表示不修改",
    ])
    ws.append(["460-00-123456-1", "100", "12345", "38400"])
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 4)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _guess_freq_band(rat: str, earfcn: Any) -> str:
    """粗略估计频段,便于分组/着色"""
    if earfcn is None:
        return "未知"
    try:
        n = int(earfcn)
    except (ValueError, TypeError):
        return "未知"
    if rat == "LTE":
        if 0 <= n <= 599: return "B1/B2/B3"
        if 1200 <= n <= 1949: return "B1/B3"
        if 2750 <= n <= 3449: return "B7"
        if 3700 <= n <= 4149: return "B8"
        if 6150 <= n <= 6449: return "B20"
        if 9210 <= n <= 9659: return "B38/B39/B40/B41"
        return f"B?({n})"
    if rat == "NR":
        if n < 600000: return "Sub6-Low"
        if n < 2425000: return "Sub6-Mid"
        return "mmWave"


def finalize_manual_cell(cell: Dict[str, Any]) -> Tuple[Dict[str, Any], List[str]]:
    """
    单条工参（界面新增/编辑）校验与派生字段，与 Excel 导入后处理一致。
    返回 (规范化后的小区, 错误列表)；错误非空时不应写入 STATE。
    """
    import copy

    cell = copy.deepcopy(cell)

    if cell.get("azimuth") is None:
        cell["azimuth"] = 0.0
    if cell.get("pci") is None:
        cell["pci"] = -1
        cell["pci_missing"] = True
    else:
        cell["pci_missing"] = False

    if cell.get("site_type") is None or cell.get("site_type") == "":
        cell["site_type"] = "陆地"

    raw_npt = cell.get("nbr_plan_types")
    if raw_npt:
        if isinstance(raw_npt, list):
            cell["nbr_plan_types"] = [str(s).strip() for s in raw_npt if str(s).strip()]
        else:
            cell["nbr_plan_types"] = [
                s.strip() for s in re.split(r"[|,]", str(raw_npt)) if s.strip()
            ]
    else:
        cell["nbr_plan_types"] = None

    cell["plan_site_type"] = to_plan_site_type(
        cell.get("plan_site_type") or cell.get("site_type")
    )

    for f in ("n_sectors", "base_azimuth", "locked", "neighbors_json", "updated_at"):
        if f not in cell or cell.get(f) is None:
            if f == "n_sectors":
                cell[f] = 1
            elif f == "base_azimuth":
                cell[f] = 0.0
            elif f == "locked":
                cell[f] = False
            elif f == "neighbors_json":
                cell[f] = "[]"
            elif f == "updated_at":
                cell[f] = None

    errs = _validate_cell(cell)
    if errs:
        return cell, errs

    freq_raw = cell.get("freq_band_raw")
    freq_band = normalize_freq_band(cell["rat"], freq_raw)
    if not freq_band and cell.get("earfcn") is not None:
        freq_band = _guess_freq_band(cell["rat"], cell.get("earfcn"))
    cell["freq_band"] = freq_band or "默认"

    enrich_cell_with_sector(cell)

    phy = cell.get("phy_name")
    if phy:
        cell["site_name"] = phy
    else:
        nm = cell.get("name") or ""
        cell["site_name"] = nm.rsplit("-", 1)[0] if "-" in nm else nm

    if "neighbors" not in cell or cell["neighbors"] is None:
        cell["neighbors"] = []

    return cell, []
    return "未知"