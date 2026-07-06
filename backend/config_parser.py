"""
网管配置Excel解析模块

解析RANCM导出的配置xlsx文件：
- 每个sheet前5行是元信息（列名、中文名、类型、描述、主键标记）
- 第6行开始是数据
- 根据配置只导入指定的sheet和列
"""
from __future__ import annotations

import io
import logging
import re
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from config_imports import (
    get_enabled_sheets,
    get_sheet_columns,
    is_sheet_enabled,
    resolve_unique_key_columns,
)

logger = logging.getLogger(__name__)

__all__ = ["parse_config_excel", "parse_single_sheet", "apply_sheet_unique_keys"]


def apply_sheet_unique_keys(
    sheet_name: str,
    column_map: Dict[str, str],
    sheet_data: Dict[str, Any],
) -> Dict[str, Any]:
    """根据 config.yaml unique_keys 或 Excel 主键标记确定 pk_columns，并同步 columns[].is_pk"""
    pk_columns, warnings = resolve_unique_key_columns(
        sheet_name, column_map, sheet_data.get("pk_columns") or []
    )
    for w in warnings:
        logger.warning(f"sheet '{sheet_name}': {w}")
    sheet_data["pk_columns"] = pk_columns
    sheet_data["unique_key_warnings"] = warnings
    pk_set = set(pk_columns)
    for col in sheet_data.get("columns") or []:
        col["is_pk"] = col.get("name") in pk_set
    return sheet_data


def _parse_type_str(type_str: str) -> str:
    """
    解析类型定义字符串，推断SQLite类型

    Examples:
        "string" -> "TEXT"
        "string:length[0..80]" -> "TEXT"
        "long:[0..16383]" -> "INTEGER"
        "double:[0..100000]" -> "REAL"
        "boolean" -> "INTEGER" (0/1)
        "enum:..." -> "TEXT"
    """
    if not type_str:
        return "TEXT"

    type_str = type_str.strip().lower()

    if type_str.startswith("long") or type_str.startswith("int"):
        return "INTEGER"
    if type_str.startswith("double") or type_str.startswith("float") or type_str.startswith("real"):
        return "REAL"
    if type_str.startswith("boolean"):
        return "INTEGER"
    # string, enum, stringArray, longArray 都用 TEXT
    return "TEXT"


def _parse_pk_mark(mark_str: str) -> Tuple[bool, bool]:
    """
    解析主键/必填标记

    Returns:
        (is_primary_key, is_required)

    Examples:
        "Primary Key" -> (True, False)
        "R" -> (False, True)
        "R-C" -> (False, True)
        "M" -> (False, False)  # M 表示可修改，非必填
        "--" -> (False, False)
    """
    if not mark_str:
        return False, False

    mark_str = str(mark_str).strip()
    is_pk = "Primary Key" in mark_str
    is_required = mark_str in ("R", "R-C") or "R" in mark_str.split(",")

    return is_pk, is_required


def _coerce_value(type_str: str, value: Any) -> Any:
    """
    根据类型定义转换值

    - INTEGER: 尝试转为int
    - REAL: 尝试转为float
    - TEXT: 保持字符串，处理枚举值
    """
    sql_type = _parse_type_str(type_str)

    if value is None:
        return None

    # 处理字符串中的枚举格式 "值[code]" -> 取值部分
    if isinstance(value, str):
        value = value.strip()
        # 处理 "是[1]", "否[0]" 等格式
        enum_match = re.match(r"^(.+?)\s*\[\d+\]$", value)
        if enum_match:
            # 对于枚举，通常只需要显示值，但存储时可能需要code
            # 这里保留原始字符串
            pass

    if sql_type == "INTEGER":
        try:
            # 处理 "是[1]" -> 1, "否[0]" -> 0
            if isinstance(value, str):
                enum_match = re.search(r"\[(\d+)\]$", value)
                if enum_match:
                    return int(enum_match.group(1))
                if value.lower() in ("true", "是", "yes", "1"):
                    return 1
                if value.lower() in ("false", "否", "no", "0"):
                    return 0
            return int(float(str(value)))
        except (ValueError, TypeError):
            return None

    if sql_type == "REAL":
        try:
            return float(str(value))
        except (ValueError, TypeError):
            return None

    # TEXT: 返回字符串
    return str(value) if value is not None else None


def parse_single_sheet(
    ws: openpyxl.worksheet.Worksheet,
    column_map: Dict[str, str],
) -> Dict[str, Any]:
    """
    解析单个sheet，只提取配置的列

    Args:
        ws: openpyxl worksheet对象
        column_map: {Excel列名: 数据库列名}

    Returns:
        {
            "columns": [{"name": 数据库列名, "type": SQLite类型, "is_pk": bool, "src": Excel列名}, ...],
            "rows": [{数据库列名: 值, ...}, ...],
            "row_count": int,
            "pk_columns": [主键列名列表],
        }
    """
    # 读取前5行元信息
    meta_rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
    if len(meta_rows) < 5:
        return {"columns": [], "rows": [], "row_count": 0, "pk_columns": [], "error": "sheet行数不足"}

    row1_cols = [c for c in meta_rows[0] if c is not None]  # 英文列名
    row2_cols = [c for c in meta_rows[1] if c is not None]  # 中文列名
    row3_cols = [c for c in meta_rows[2] if c is not None]  # 类型
    row4_cols = [c for c in meta_rows[3] if c is not None]  # 描述
    row5_cols = [c for c in meta_rows[4] if c is not None]  # 主键标记

    # 建立列名到索引的映射
    col_index_map: Dict[str, int] = {}
    for i, col_name in enumerate(row1_cols):
        col_name_str = str(col_name).strip() if col_name else ""
        if col_name_str:
            col_index_map[col_name_str] = i

    # 只保留配置中的列
    columns_info: List[Dict[str, Any]] = []
    pk_columns: List[str] = []

    for src_col, dst_col in column_map.items():
        if src_col not in col_index_map:
            logger.warning(f"列 '{src_col}' 在sheet中不存在，跳过")
            continue

        idx = col_index_map[src_col]
        type_str = str(row3_cols[idx] or "") if idx < len(row3_cols) else ""
        pk_mark = str(row5_cols[idx] or "") if idx < len(row5_cols) else ""
        is_pk, is_required = _parse_pk_mark(pk_mark)

        columns_info.append({
            "name": dst_col,
            "type": _parse_type_str(type_str),
            "is_pk": is_pk,
            "is_required": is_required,
            "src": src_col,
            "desc": str(row4_cols[idx] or "") if idx < len(row4_cols) else "",
        })

        if is_pk:
            pk_columns.append(dst_col)

    if not columns_info:
        return {"columns": [], "rows": [], "row_count": 0, "pk_columns": [], "error": "无匹配列"}

    # 从第6行开始读取数据
    rows_data: List[Dict[str, Any]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
        row_dict: Dict[str, Any] = {}
        has_data = False

        for col_info in columns_info:
            src_col = col_info["src"]
            dst_col = col_info["name"]
            idx = col_index_map[src_col]

            raw_value = row[idx] if idx < len(row) else None
            if raw_value is not None:
                has_data = True

            # 类型转换
            type_str = ""
            # 从元信息中获取类型
            if idx < len(row3_cols):
                type_str = str(row3_cols[idx] or "")

            converted = _coerce_value(type_str, raw_value)
            row_dict[dst_col] = converted

        if has_data:
            rows_data.append(row_dict)

    return {
        "columns": columns_info,
        "rows": rows_data,
        "row_count": len(rows_data),
        "pk_columns": pk_columns,
    }


def parse_config_excel(
    file_bytes: bytes,
    filename: str,
    sheet_names: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """
    解析配置Excel文件，只导入配置的sheet和列

    Args:
        file_bytes: Excel文件二进制内容
        filename: 文件名（用于日志）
        sheet_names: 要解析的sheet名列表，为None时使用配置中的启用sheet

    Returns:
        {
            "success": True/False,
            "error": "错误信息（如有）",
            "filename": str,
            "sheets": {
                sheet_name: {
                    "columns": [...],
                    "rows": [...],
                    "row_count": int,
                    "pk_columns": [...],
                }
            },
            "stats": {
                "total_sheets": int,
                "total_rows": int,
                "skipped_sheets": [未解析的sheet名],
            }
        }
    """
    result: Dict[str, Any] = {
        "success": False,
        "error": None,
        "filename": filename,
        "sheets": {},
        "stats": {
            "total_sheets": 0,
            "total_rows": 0,
            "skipped_sheets": [],
        },
    }

    # 确定要解析的sheet
    if sheet_names is None:
        sheet_names = get_enabled_sheets()

    if not sheet_names:
        result["error"] = "未指定要解析的sheet，且配置中无启用的sheet"
        return result

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        result["error"] = f"Excel文件读取失败: {e}"
        logger.error(f"解析配置Excel失败: {filename}, {e}")
        return result

    result["stats"]["total_sheets"] = len(sheet_names)

    for sheet_name in sheet_names:
        # 检查sheet是否存在于Excel中
        if sheet_name not in wb.sheetnames:
            result["stats"]["skipped_sheets"].append(f"{sheet_name}(不存在)")
            logger.warning(f"sheet '{sheet_name}' 在文件 '{filename}' 中不存在")
            continue

        # 获取该sheet的列配置
        column_map = get_sheet_columns(sheet_name)
        if not column_map:
            result["stats"]["skipped_sheets"].append(f"{sheet_name}(无列配置)")
            logger.warning(f"sheet '{sheet_name}' 无列配置，跳过")
            continue

        # 检查是否启用
        if not is_sheet_enabled(sheet_name):
            result["stats"]["skipped_sheets"].append(f"{sheet_name}(未启用)")
            continue

        try:
            ws = wb[sheet_name]
            sheet_data = parse_single_sheet(ws, column_map)

            if sheet_data.get("error"):
                result["stats"]["skipped_sheets"].append(f"{sheet_name}({sheet_data['error']})")
                continue

            apply_sheet_unique_keys(sheet_name, column_map, sheet_data)
            result["sheets"][sheet_name] = sheet_data
            result["stats"]["total_rows"] += sheet_data["row_count"]

            logger.info(
                f"解析sheet '{sheet_name}': {sheet_data['row_count']} 行, "
                f"{len(sheet_data['columns'])} 列"
            )

        except Exception as e:
            result["stats"]["skipped_sheets"].append(f"{sheet_name}(解析错误: {e})")
            logger.error(f"解析sheet '{sheet_name}' 失败: {e}")

    wb.close()

    # 至少成功解析一个sheet才算成功
    if result["sheets"]:
        result["success"] = True
    else:
        result["error"] = "未能成功解析任何sheet"

    return result


def list_excel_sheets(file_bytes: bytes, filename: str) -> Dict[str, Any]:
    """
    列出Excel文件中的所有sheet名称（用于预览）

    Returns:
        {
            "success": True/False,
            "error": str,
            "sheets": [sheet名列表],
            "configured": [已配置的sheet名],
            "unconfigured": [未配置的sheet名],
        }
    """
    result: Dict[str, Any] = {
        "success": False,
        "error": None,
        "sheets": [],
        "configured": [],
        "unconfigured": [],
    }

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        result["sheets"] = wb.sheetnames
        wb.close()
    except Exception as e:
        result["error"] = f"读取Excel失败: {e}"
        return result

    # 区分已配置和未配置
    enabled_sheets = get_enabled_sheets()
    all_configured = set(enabled_sheets)

    for s in result["sheets"]:
        if s in all_configured:
            result["configured"].append(s)
        else:
            result["unconfigured"].append(s)

    result["success"] = True
    return result