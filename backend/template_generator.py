"""
工参模板生成器 - 创建可下载的标准 Excel 模板
- 4G LTE 工参模板 (1 sheet)
- 5G NR 工参模板 (1 sheet)
表头标注 [必填]/[可选] 与 [枚举]/[开放]；字段说明 sheet 与解析器 data_parser 一致
"""
from __future__ import annotations

from io import BytesIO
from typing import Dict, List, Tuple

import pandas as pd

# ──────────────────────────────────────────────
# 批量规划模板列（逻辑名 → 必填、值类型、说明）
# 值类型: 枚举 = 须从给定取值中选；开放 = 自由填写（可有推荐值）
# ──────────────────────────────────────────────
BATCH_COLUMN_META: Dict[str, Dict[str, str]] = {
    "CGI": {
        "required": "必填",
        "value_kind": "开放",
        "enum_hint": "",
        "doc": "小区全球标识，格式 MCC-MNC-基站ID-小区ID（如 460-00-123456-1）",
    },
    "小区名称": {
        "required": "必填",
        "value_kind": "开放",
        "enum_hint": "",
        "doc": "中文/英文小区名",
    },
    "网络制式": {
        "required": "必填",
        "value_kind": "枚举",
        "enum_hint": "4G | 5G",
        "doc": "填写 4G 或 5G（亦支持 LTE/NR）",
    },
    "经度": {
        "required": "必填",
        "value_kind": "开放",
        "enum_hint": "",
        "doc": "WGS84，度，范围 -180 ~ 180",
    },
    "纬度": {
        "required": "必填",
        "value_kind": "开放",
        "enum_hint": "",
        "doc": "WGS84，度，范围 -90 ~ 90",
    },
    "方位角": {
        "required": "必填",
        "value_kind": "开放",
        "enum_hint": "",
        "doc": "正北为 0° 顺时针，范围 0 ~ 360",
    },
    "覆盖半径": {
        "required": "可选",
        "value_kind": "开放",
        "enum_hint": "",
        "doc": "单位米；留空则按站点类型+制式+频段自动映射（推荐）",
    },
    "物理小区识别码": {
        "required": "可选",
        "value_kind": "开放",
        "enum_hint": "留空=规划分配；已有站填 0-503；锁定=是 时须填原 PCI",
        "doc": "4G PCI。批量新建建议留空，规划结果在导出 xlsx 的 PCI 列",
    },
    "PCI": {
        "required": "可选",
        "value_kind": "开放",
        "enum_hint": "留空=规划分配；已有站填 0-1007；锁定=是 时须填原 PCI",
        "doc": "5G PCI。批量新建建议留空，规划结果在导出 xlsx 的 PCI 列",
    },
    "TAC": {
        "required": "可选",
        "value_kind": "开放",
        "enum_hint": "",
        "doc": "跟踪区码",
    },
    "详细使用频段": {
        "required": "必填",
        "value_kind": "开放",
        "enum_hint": "见「扇区参数映射表」频段描述",
        "doc": "频段描述，系统自动识别（FDD900/FDD1800/F/D/A/700M/2.6G/4.9G 等）",
    },
    "站点类型": {
        "required": "可选",
        "value_kind": "枚举",
        "enum_hint": "陆地 | 室内 | 近海 | 微站",
        "doc": "默认陆地（宏站场景）",
    },
    "所属基站名称": {
        "required": "可选",
        "value_kind": "开放",
        "enum_hint": "",
        "doc": "站点名，用于分组/着色",
    },
    "扇区数": {
        "required": "可选",
        "value_kind": "枚举",
        "enum_hint": "1 | 2 | 3 | 4 | 5 | 6",
        "doc": "默认 1；多扇区时方位角=基方位角+i×360/扇区数",
    },
    "基方位角": {
        "required": "可选",
        "value_kind": "开放",
        "enum_hint": "",
        "doc": "0-360，默认 0",
    },
    "规划类型": {
        "required": "可选",
        "value_kind": "枚举",
        "enum_hint": "宏站 | 微站 | 室分",
        "doc": "不填则按「站点类型」自动归类",
    },
    "邻区规划": {
        "required": "可选",
        "value_kind": "枚举",
        "enum_hint": "4G_4G | 4G_5G | 5G_4G | 5G_5G（多选用 | 连接）",
        "doc": "本行参与哪些邻区关系规划",
    },
    "邻区得分阈值": {
        "required": "可选",
        "value_kind": "开放",
        "enum_hint": "0 ~ 1，留空=0.5",
        "doc": "本小区邻区规划最低得分；低于阈值的候选邻区丢弃（第一圈复用半径内仍强制保留）",
    },
    "锁定": {
        "required": "可选",
        "value_kind": "枚举",
        "enum_hint": "是 | 否",
        "doc": "是=不参与 PCI 重分配，保留原 PCI",
    },
}

# 工参 sheet 第 2 行「取值/范围」说明（与表头列一一对应；上传解析会自动跳过此行）
BATCH_RANGE_ROW: Dict[str, str] = {
    "CGI": "MCC-MNC-基站ID-小区ID，如 460-00-123456-1",
    "小区名称": "任意文本",
    "网络制式": "4G 或 5G（亦支持 LTE / NR 等别名）",
    "经度": "-180 ~ 180（WGS84）",
    "纬度": "-90 ~ 90（WGS84）",
    "方位角": "0 ~ 360（≥360 会模 360）",
    "覆盖半径": "米，>0；留空则按站型+制式+频段映射",
    "物理小区识别码": "留空=规划分配；4G：0–503；锁定=是 时填原 PCI",
    "PCI": "留空=规划分配；5G：0–1007；锁定=是 时填原 PCI",
    "TAC": "整数",
    "详细使用频段": "频段描述（FDD900/F/D/A、Band3、700M/2.6G/4.9G 等），见「扇区参数映射表」",
    "站点类型": "陆地 | 室内 | 近海 | 微站",
    "所属基站名称": "站点名",
    "扇区数": "1 | 2 | 3 | 4 | 5 | 6，默认 1",
    "基方位角": "0–360，默认 0",
    "规划类型": "宏站 | 微站 | 室分（不填则由站点类型推断）",
    "邻区规划": "4G_4G | 4G_5G | 5G_4G | 5G_5G，多选用 | 连接",
    "邻区得分阈值": "0 ~ 1 小数，如 0.5；留空默认 0.5",
    "锁定": "是 | 否（是=不参与 PCI 重分配）",
}

LTE_COLUMNS_BASE = [
    "CGI", "小区名称", "网络制式", "经度", "纬度",
    "方位角", "覆盖半径", "物理小区识别码", "TAC",
    "详细使用频段", "站点类型", "所属基站名称",
    "扇区数", "基方位角", "规划类型", "邻区规划", "邻区得分阈值", "锁定",
]

NR_COLUMNS_BASE = [
    "CGI", "小区名称", "网络制式", "经度", "纬度",
    "方位角", "覆盖半径", "PCI", "TAC",
    "详细使用频段", "站点类型", "所属基站名称",
    "扇区数", "基方位角", "规划类型", "邻区规划", "邻区得分阈值", "锁定",
]

# 兼容旧引用
LTE_COLUMNS = LTE_COLUMNS_BASE
NR_COLUMNS = NR_COLUMNS_BASE


def _header_label(logical_name: str) -> str:
    meta = BATCH_COLUMN_META.get(logical_name, {})
    req = meta.get("required", "可选")
    kind = meta.get("value_kind", "开放")
    return f"{logical_name}[{req}][{kind}]"


def _headers_for_columns(base_columns: List[str]) -> List[str]:
    return [_header_label(c) for c in base_columns]


def _range_row_values(base_columns: List[str]) -> List[str]:
    return [BATCH_RANGE_ROW.get(c, "") for c in base_columns]


def _logical_name_from_header(header: str) -> str:
    raw = str(header or "")
    for base in BATCH_COLUMN_META:
        if raw.startswith(base):
            return base
    return raw.split("[")[0].strip()


def _style_template_sheet(ws, base_columns: List[str]) -> None:
    """表头配色：必填=红字，可选=灰字；枚举列浅黄底，开放列浅蓝底"""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    fill_enum = PatternFill("solid", fgColor="FFF8E6")
    fill_open = PatternFill("solid", fgColor="E8F4FC")
    font_req = Font(bold=True, color="C00000", size=10)
    font_opt = Font(bold=True, color="595959", size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    n_cols = len(base_columns)

    for col_idx, logical in enumerate(base_columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        meta = BATCH_COLUMN_META.get(logical, {})
        cell.font = font_req if meta.get("required") == "必填" else font_opt
        cell.fill = fill_enum if meta.get("value_kind") == "枚举" else fill_open
        cell.alignment = align
        letter = get_column_letter(col_idx)
        wide_cols = {"小区名称", "详细使用频段", "所属基站名称", "邻区规划"}
        ws.column_dimensions[letter].width = 24 if logical in wide_cols else 16

    ws.row_dimensions[1].height = 40


def _style_range_row(ws, base_columns: List[str]) -> None:
    """第 2 行：各列取值/范围说明"""
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor="F0F2F5")
    font = Font(size=9, color="404040", italic=True)
    align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for col_idx, logical in enumerate(base_columns, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = BATCH_RANGE_ROW.get(logical, "")
        cell.font = font
        cell.fill = fill
        cell.alignment = align
    ws.row_dimensions[2].height = 52


def _insert_range_row_after_header(ws, base_columns: List[str]) -> None:
    """在表头下插入取值说明行（示例数据从第 3 行起）"""
    ws.insert_rows(2)
    _style_range_row(ws, base_columns)


def _sample_rows():
    """阳江市区附近示例数据,确保每种频段至少一个示例"""
    # 阳江市政府附近 21.86°N, 111.95°E
    samples = [
        # 5G 700M
        {
            "CGI": "460-00-123456-101",
            "小区名称": "阳江市政府_5G_700M_1",
            "网络制式": "5G",
            "经度": 111.952,
            "纬度": 21.859,
            "方位角": 0,
            "覆盖半径": 50,
            "PCI": 100,
            "物理小区识别码": 100,
            "TAC": 41001,
            "详细使用频段": "n28 700MHz",
            "站点类型": "陆地",
            "所属基站名称": "阳江市政府",
            "扇区数": 3, "基方位角": 0, "规划类型": "宏站",
            "邻区规划": "5G_4G|5G_5G", "锁定": "否",
        },
        # 5G 2.6G
        {
            "CGI": "460-00-123456-102",
            "小区名称": "阳江市政府_5G_2.6G_1",
            "网络制式": "5G",
            "经度": 111.952,
            "纬度": 21.859,
            "方位角": 120,
            "覆盖半径": 40,
            "PCI": 200,
            "物理小区识别码": 200,
            "TAC": 41001,
            "详细使用频段": "n41 2.6GHz",
            "站点类型": "陆地",
            "所属基站名称": "阳江市政府",
            "扇区数": 3, "基方位角": 120, "规划类型": "宏站",
            "邻区规划": "5G_4G|5G_5G", "锁定": "否",
        },
        # 5G 4.9G
        {
            "CGI": "460-00-123456-103",
            "小区名称": "阳江市政府_5G_4.9G_1",
            "网络制式": "5G",
            "经度": 111.952,
            "纬度": 21.859,
            "方位角": 240,
            "覆盖半径": 30,
            "PCI": 300,
            "物理小区识别码": 300,
            "TAC": 41001,
            "详细使用频段": "n77 4.9GHz",
            "站点类型": "陆地",
            "所属基站名称": "阳江市政府",
            "扇区数": 3, "基方位角": 240, "规划类型": "宏站",
            "邻区规划": "5G_4G|5G_5G", "锁定": "否",
        },
        # 4G FDD900
        {
            "CGI": "460-00-123450-201",
            "小区名称": "阳江市政府_4G_FDD900_1",
            "网络制式": "4G",
            "经度": 111.952,
            "纬度": 21.859,
            "方位角": 0,
            "覆盖半径": 47,
            "PCI": 50,
            "物理小区识别码": 50,
            "TAC": 41001,
            "详细使用频段": "Band8 900MHz FDD",
            "站点类型": "陆地",
            "所属基站名称": "阳江市政府",
            "扇区数": 3, "基方位角": 0, "规划类型": "宏站",
            "邻区规划": "4G_4G|4G_5G", "锁定": "否",
        },
        # 4G FDD1800
        {
            "CGI": "460-00-123450-202",
            "小区名称": "阳江市政府_4G_FDD1800_1",
            "网络制式": "4G",
            "经度": 111.952,
            "纬度": 21.859,
            "方位角": 120,
            "覆盖半径": 43,
            "PCI": 150,
            "物理小区识别码": 150,
            "TAC": 41001,
            "详细使用频段": "Band3 1800MHz FDD",
            "站点类型": "陆地",
            "所属基站名称": "阳江市政府",
            "扇区数": 3, "基方位角": 120, "规划类型": "宏站",
            "邻区规划": "4G_4G|4G_5G", "锁定": "否",
        },
        # 4G F 频段
        {
            "CGI": "460-00-123450-203",
            "小区名称": "阳江市政府_4G_F频_1",
            "网络制式": "4G",
            "经度": 111.952,
            "纬度": 21.859,
            "方位角": 240,
            "覆盖半径": 39,
            "PCI": 250,
            "物理小区识别码": 250,
            "TAC": 41001,
            "详细使用频段": "F频段 1885MHz",
            "站点类型": "陆地",
            "所属基站名称": "阳江市政府",
            "扇区数": 3, "基方位角": 240, "规划类型": "宏站",
            "邻区规划": "4G_4G|4G_5G", "锁定": "否",
        },
        # 4G D 频段
        {
            "CGI": "460-00-123450-204",
            "小区名称": "阳江市政府_4G_D频_1",
            "网络制式": "4G",
            "经度": 111.952,
            "纬度": 21.859,
            "方位角": 30,
            "覆盖半径": 42,
            "PCI": 350,
            "物理小区识别码": 350,
            "TAC": 41001,
            "详细使用频段": "D频段 2575MHz",
            "站点类型": "陆地",
            "所属基站名称": "阳江市政府",
            "扇区数": 3, "基方位角": 30, "规划类型": "宏站",
            "邻区规划": "4G_4G|4G_5G", "锁定": "否",
        },
        # 4G A 频段
        {
            "CGI": "460-00-123450-205",
            "小区名称": "阳江市政府_4G_A频_1",
            "网络制式": "4G",
            "经度": 111.952,
            "纬度": 21.859,
            "方位角": 150,
            "覆盖半径": 38,
            "PCI": 450,
            "物理小区识别码": 450,
            "TAC": 41001,
            "详细使用频段": "A频段 2010MHz",
            "站点类型": "陆地",
            "所属基站名称": "阳江市政府",
            "扇区数": 3, "基方位角": 150, "规划类型": "宏站",
            "邻区规划": "4G_4G|4G_5G", "锁定": "否",
        },
        # 室分
        {
            "CGI": "460-00-123450-301",
            "小区名称": "阳江市政府大楼_室分_1",
            "网络制式": "4G",
            "经度": 111.952,
            "纬度": 21.859,
            "方位角": 0,
            "覆盖半径": 30,
            "PCI": 80,
            "物理小区识别码": 80,
            "TAC": 41001,
            "详细使用频段": "Band3 1800MHz",
            "站点类型": "室内",
            "所属基站名称": "阳江市政府大楼_室内分布",
            "扇区数": 1, "基方位角": 0, "规划类型": "室分",
            "邻区规划": "4G_4G", "锁定": "否",
        },
        # 微站 (新增)
        {
            "CGI": "460-00-123450-401",
            "小区名称": "阳江市政府_微站_1",
            "网络制式": "4G",
            "经度": 111.953,
            "纬度": 21.860,
            "方位角": 0,
            "覆盖半径": 40,
            "PCI": 60,
            "物理小区识别码": 60,
            "TAC": 41001,
            "详细使用频段": "F频段 1885MHz",
            "站点类型": "微站",
            "所属基站名称": "阳江市政府_微站",
            "扇区数": 1, "基方位角": 0, "规划类型": "微站",
            "邻区规划": "4G_4G|4G_5G", "锁定": "否",
        },
    ]
    return samples


def _build_field_doc_rows(columns: List[str]) -> List[Tuple[str, str, str, str, str]]:
    """字段说明 sheet 行，列顺序与对应工参 sheet 一致（4G 或 5G）"""
    rows: List[Tuple[str, str, str, str, str]] = []
    for logical in columns:
        meta = BATCH_COLUMN_META.get(logical)
        if not meta:
            continue
        rows.append((
            logical,
            meta["required"],
            meta["value_kind"],
            meta.get("enum_hint", "") or "—",
            meta["doc"],
        ))
    return rows


def _legend_rows(rat: str) -> List[Tuple[str, str, str]]:
    pci_col = "PCI" if rat == "5G" else ("物理小区识别码" if rat == "4G" else "PCI（5G）/ 物理小区识别码（4G）")
    rows: List[Tuple[str, str, str]] = [
        ("[必填]", "上传解析缺少该列或该格为空会导致该行无效", ""),
        ("[可选]", "可留空，系统按默认值或自动映射填补", ""),
        ("[枚举]", "须从「枚举/取值说明」中选取；多选字段用 | 连接", ""),
        ("[开放]", "自由填写；可参考示例或「扇区参数映射表」", ""),
        ("", "", ""),
        (
            "本工作簿制式",
            {"4G": "仅 4G LTE", "5G": "仅 5G NR", "both": "4G + 5G 双 sheet"}.get(rat, rat),
            "",
        ),
        (
            "与 data_parser 一致",
            "必填列: CGI、小区名称、网络制式、经度、纬度、方位角",
            "",
        ),
        (
            "PCI 列（可选）",
            f"{pci_col} 可不填，由规划写入导出表；锁定=是 时填写原 PCI 以保留",
            "",
        ),
        ("", "强烈建议填: 详细使用频段（用于扇区半径/波束映射）", ""),
        (
            "第 2 行",
            "各列「取值/范围」说明行；上传批量规划时会自动跳过，勿删表头",
            "",
        ),
    ]
    if rat in ("5G", "both"):
        rows.append((
            "5G 工参 sheet",
            "PCI[可选][开放]：新建留空由规划分配（0–1007）；锁定=是 时填原 PCI；结果见导出 xlsx",
            "",
        ))
    if rat in ("4G", "both"):
        rows.append((
            "4G 工参 sheet",
            "物理小区识别码[可选][开放]：新建留空由规划分配（0–503）；锁定=是 时填原 PCI；结果见导出 xlsx",
            "",
        ))
    return rows

# 规划类型与 PCI 距离阈值
PLAN_TYPE_DOC = [
    ("宏站", "macro",  700,  5000, "覆盖距离700m, 5km无PCI相同规划"),
    ("微站", "micro",  200,  3000, "覆盖距离200m, 3km无PCI相同规划"),
    ("室分", "indoor", 100,  2000, "覆盖距离100m, 2km无PCI相同规划"),
]

# 邻区规划类型枚举
NBR_TYPE_DOC = [
    ("4G_4G", "LTE → LTE", "同制式4G邻区"),
    ("4G_5G", "LTE → NR",  "4G到5G邻区"),
    ("5G_4G", "NR → LTE",  "5G到4G邻区(锚点)"),
    ("5G_5G", "NR → NR",   "同制式5G邻区"),
]


def generate_template(rat: str = "both") -> bytes:
    """
    生成工参模板 Excel
    rat: '4G' / '5G' / 'both'
  表头含 [必填]/[可选] 与 [枚举]/[开放] 标识
    """
    output = BytesIO()
    samples = _sample_rows()
    doc_rows: List[Tuple[str, str, str, str, str]] = []
    if rat in ("4G", "both"):
        if rat == "both":
            doc_rows.append(("── 4G工参模板列 ──", "—", "—", "—", "见 sheet「4G工参模板」表头"))
        doc_rows.extend(_build_field_doc_rows(LTE_COLUMNS_BASE))
    if rat in ("5G", "both"):
        if rat == "both":
            doc_rows.append(("── 5G工参模板列 ──", "—", "—", "—", "见 sheet「5G工参模板」表头"))
        doc_rows.extend(_build_field_doc_rows(NR_COLUMNS_BASE))

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if rat in ("4G", "both"):
            df_4g = pd.DataFrame(
                [s for s in samples if s["网络制式"] == "4G"],
                columns=LTE_COLUMNS_BASE,
            )
            df_4g.columns = _headers_for_columns(LTE_COLUMNS_BASE)
            df_4g.to_excel(writer, sheet_name="4G工参模板", index=False)
            _style_template_sheet(writer.sheets["4G工参模板"], LTE_COLUMNS_BASE)
            _insert_range_row_after_header(writer.sheets["4G工参模板"], LTE_COLUMNS_BASE)

        if rat in ("5G", "both"):
            df_5g = pd.DataFrame(
                [s for s in samples if s["网络制式"] == "5G"],
                columns=NR_COLUMNS_BASE,
            )
            df_5g.columns = _headers_for_columns(NR_COLUMNS_BASE)
            df_5g.to_excel(writer, sheet_name="5G工参模板", index=False)
            _style_template_sheet(writer.sheets["5G工参模板"], NR_COLUMNS_BASE)
            _insert_range_row_after_header(writer.sheets["5G工参模板"], NR_COLUMNS_BASE)

        pd.DataFrame(_legend_rows(rat), columns=["标识", "含义", "备注"]).to_excel(
            writer, sheet_name="表头图例", index=False
        )

        # 字段说明 sheet
        df_doc = pd.DataFrame(
            doc_rows,
            columns=["字段名", "是否必填", "值类型", "枚举/取值说明", "详细说明"],
        )
        df_doc.to_excel(writer, sheet_name="字段说明", index=False)

        # 频段映射表 (供用户参考)
        freq_map = [
            ("5G",  "700M",  40, 50, "n28 / 700MHz"),
            ("5G",  "2.6G",  65, 40, "n41 / 2.6GHz"),
            ("5G",  "4.9G",  70, 30, "n77 / 4.9GHz"),
            ("4G",  "FDD900",  30, 47, "Band8 900MHz"),
            ("4G",  "FDD1800", 50, 43, "Band3 1800MHz"),
            ("4G",  "F",       45, 39, "F频段 (1885MHz)"),
            ("4G",  "D",       60, 42, "D频段 (2575MHz)"),
            ("4G",  "A",       55, 38, "A频段 (2010MHz)"),
            ("室内", "—",      359, 30, "室分全向"),
            ("微站", "—",      40, 40, "微站(40°/40m)"),
        ]
        df_freq = pd.DataFrame(freq_map, columns=["制式", "频段", "波瓣全角(°)", "覆盖半径(m)", "频段描述"])
        df_freq.to_excel(writer, sheet_name="扇区参数映射表", index=False)

        # 规划类型说明 sheet
        df_ptype = pd.DataFrame(PLAN_TYPE_DOC, columns=["中文标签", "代码", "safe_distance_m", "same_pci_min_m", "说明"])
        df_ptype.to_excel(writer, sheet_name="规划类型说明", index=False)

        # 邻区规划类型说明 sheet
        df_ntype = pd.DataFrame(NBR_TYPE_DOC, columns=["枚举值", "方向", "说明"])
        df_ntype.to_excel(writer, sheet_name="邻区规划类型说明", index=False)

    return output.getvalue()